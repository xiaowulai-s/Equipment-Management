# -*- coding: utf-8 -*-
"""Helpers for exporting runtime and historical data."""

from __future__ import annotations

import csv
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


class DataExporter:
    """Utility methods for exporting data to common file formats."""

    @staticmethod
    def export_to_csv(
        data: List[Dict[str, Any]], file_path: str, include_headers: bool = True, encoding: str = "utf-8-sig"
    ) -> bool:
        """Write dictionaries to a CSV file."""
        if not data:
            return False

        try:
            path = Path(file_path)
            path.parent.mkdir(parents=True, exist_ok=True)
            fieldnames = list(data[0].keys())

            with open(path, "w", newline="", encoding=encoding) as csv_file:
                writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
                if include_headers:
                    writer.writeheader()
                writer.writerows(data)
            return True
        except Exception:
            logger.exception("导出 CSV 失败: %s", file_path)
            return False

    @staticmethod
    def export_device_data_to_csv(
        devices_data: List[Dict[str, Any]], file_path: str, include_headers: bool = True
    ) -> bool:
        """Export normalized device payloads to CSV."""
        if not devices_data:
            return False

        export_rows = []
        for device in devices_data:
            row = {
                "设备 ID": device.get("device_id", ""),
                "设备名称": device.get("name", ""),
                "设备类型": device.get("device_type") or device.get("type", ""),
                "状态": device.get("status_text", device.get("status", "")),
                "主机地址": device.get("host") or device.get("ip") or device.get("ip_address", ""),
                "端口": device.get("port", ""),
                "单元 ID": device.get("unit_id", device.get("slave_id", "")),
            }

            current_data = device.get("current_data", {})
            for parameter_name, parameter_info in current_data.items():
                if not isinstance(parameter_info, dict):
                    row[parameter_name] = parameter_info
                    continue
                unit = parameter_info.get("unit", "")
                column_name = f"{parameter_name}({unit})" if unit else parameter_name
                row[column_name] = parameter_info.get("value", "")

            row["导出时间"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            export_rows.append(row)

        return DataExporter.export_to_csv(export_rows, file_path, include_headers)

    @staticmethod
    def export_alarm_history_to_csv(alarms: List[Dict[str, Any]], file_path: str, include_headers: bool = True) -> bool:
        """Export alarm history to CSV."""
        if not alarms:
            return False

        export_rows = []
        for alarm in alarms:
            export_rows.append(
                {
                    "报警 ID": alarm.get("alarm_id", alarm.get("id", "")),
                    "设备 ID": alarm.get("device_id", ""),
                    "参数": alarm.get("parameter", ""),
                    "报警类型": alarm.get("alarm_type", ""),
                    "级别": alarm.get("level_name", alarm.get("level", "")),
                    "值": alarm.get("value", ""),
                    "高阈值": alarm.get("threshold_high", ""),
                    "低阈值": alarm.get("threshold_low", ""),
                    "描述": alarm.get("description", ""),
                    "时间": alarm.get("timestamp", ""),
                    "已确认": "是" if alarm.get("acknowledged") else "否",
                    "已清除": "是" if alarm.get("cleared") else "否",
                }
            )

        return DataExporter.export_to_csv(export_rows, file_path, include_headers)

    @staticmethod
    def export_to_json(data: Any, file_path: str, indent: int = 2) -> bool:
        """Write data to JSON."""
        try:
            path = Path(file_path)
            path.parent.mkdir(parents=True, exist_ok=True)
            with open(path, "w", encoding="utf-8") as file:
                json.dump(data, file, ensure_ascii=False, indent=indent, default=str)
            return True
        except Exception:
            logger.exception("导出 JSON 失败: %s", file_path)
            return False

    @staticmethod
    def export_to_excel(
        data: List[Dict[str, Any]], file_path: str, sheet_name: str = "数据", include_headers: bool = True
    ) -> bool:
        """Write data to an Excel workbook using openpyxl when available."""
        if not data:
            return False

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            path = Path(file_path)
            path.parent.mkdir(parents=True, exist_ok=True)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name

            headers = list(data[0].keys())
            if include_headers:
                for column, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=column, value=header)
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                for column in range(1, len(headers) + 1):
                    cell = worksheet.cell(row=1, column=column)
                    cell.font = header_font
                    cell.fill = header_fill

            start_row = 2 if include_headers else 1
            for row_index, row_data in enumerate(data, start_row):
                for column_index, value in enumerate(row_data.values(), 1):
                    worksheet.cell(row=row_index, column=column_index, value=value)

            for column_cells in worksheet.columns:
                max_length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
                worksheet.column_dimensions[column_cells[0].column_letter].width = min((max_length + 2) * 1.2, 50)

            workbook.save(path)
            return True
        except ImportError:
            logger.error("导出 Excel 失败: 缺少 openpyxl 依赖")
            return False
        except Exception:
            logger.exception("导出 Excel 失败: %s", file_path)
            return False

    @staticmethod
    def generate_export_filename(prefix: str = "export", extension: str = "csv") -> str:
        """Generate a timestamped export filename."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{extension}"
