# -*- coding: utf-8 -*-
"""
工业设备管理系统 - 数据持久化模块
使用SQLite存储历史数据和配置信息
"""

import sqlite3
import os
import logging
import json
import csv
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime, timedelta
from pathlib import Path

# 使用统一日志配置
from logging_config import get_logger
logger = get_logger(__name__)

# 尝试导入Excel相关库
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    logger.warning("未安装openpyxl，Excel导出功能将不可用")


class DatabaseManager:
    """
    数据库管理器
    负责创建数据库表、存储和查询数据
    """

    def __init__(self, db_path: str = "data/equipment_management.db"):
        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(exist_ok=True)
        self.conn = None
        self.cursor = None
        self._init_db()

    def _init_db(self):
        """初始化数据库连接和表结构"""
        try:
            # 连接数据库
            self.conn = sqlite3.connect(
                str(self.db_path),
                detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES,
                check_same_thread=False
            )
            self.cursor = self.conn.cursor()

            # 启用外键约束
            self.cursor.execute("PRAGMA foreign_keys = ON")

            # 创建表
            self._create_tables()
            logger.info(f"数据库初始化完成: {self.db_path}")

        except Exception as e:
            logger.error(f"数据库初始化失败: {e}")
            raise

    def _create_tables(self):
        """创建数据库表结构"""
        try:
            # 设备信息表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS devices (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    device_id TEXT UNIQUE NOT NULL,
                    name TEXT NOT NULL,
                    ip_address TEXT NOT NULL,
                    port INTEGER NOT NULL DEFAULT 502,
                    slave_id INTEGER NOT NULL DEFAULT 1,
                    product_id TEXT,
                    group_name TEXT,
                    description TEXT,
                    status INTEGER NOT NULL DEFAULT 0,
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            # 历史数据表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS device_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    device_id TEXT NOT NULL,
                    timestamp TIMESTAMP NOT NULL,
                    register_address INTEGER NOT NULL,
                    register_name TEXT NOT NULL,
                    value REAL NOT NULL,
                    unit TEXT,
                    status INTEGER NOT NULL DEFAULT 0,
                    FOREIGN KEY (device_id) REFERENCES devices(device_id)
                        ON DELETE CASCADE
                )
            ''')

            # 创建历史数据索引
            self.cursor.execute('''
                CREATE INDEX IF NOT EXISTS idx_device_history 
                ON device_history (device_id, register_address, timestamp DESC)
            ''')

            # 报警记录表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS alarms (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    device_id TEXT NOT NULL,
                    alarm_id TEXT NOT NULL,
                    alarm_name TEXT NOT NULL,
                    register_address INTEGER NOT NULL,
                    register_name TEXT NOT NULL,
                    trigger_value REAL NOT NULL,
                    threshold REAL NOT NULL,
                    condition TEXT NOT NULL,
                    severity TEXT NOT NULL,
                    message TEXT NOT NULL,
                    status TEXT NOT NULL DEFAULT 'active',
                    timestamp TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    cleared_at TIMESTAMP,
                    FOREIGN KEY (device_id) REFERENCES devices(device_id)
                        ON DELETE CASCADE
                )
            ''')

            # 创建报警记录索引
            self.cursor.execute('''
                CREATE INDEX IF NOT EXISTS idx_alarms 
                ON alarms (device_id, status, timestamp DESC)
            ''')

            # 通信日志表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS communication_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    device_id TEXT,
                    host TEXT NOT NULL,
                    port INTEGER NOT NULL,
                    direction TEXT NOT NULL,
                    function_code INTEGER,
                    address INTEGER,
                    count INTEGER,
                    data TEXT,
                    raw_data TEXT NOT NULL,
                    timestamp TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    success INTEGER NOT NULL DEFAULT 1,
                    error_message TEXT
                )
            ''')

            # 创建通信日志索引
            self.cursor.execute('''
                CREATE INDEX IF NOT EXISTS idx_communication_logs 
                ON communication_logs (device_id, timestamp DESC)
            ''')

            # 系统配置表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS system_config (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    key TEXT UNIQUE NOT NULL,
                    value TEXT NOT NULL,
                    description TEXT,
                    updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            # 提交事务
            self.conn.commit()

        except Exception as e:
            logger.error(f"创建表失败: {e}")
            self.conn.rollback()
            raise

    def close(self):
        """关闭数据库连接"""
        if self.conn:
            self.conn.close()
            logger.info("数据库连接已关闭")

    def add_device(self, device_id: str, name: str, ip_address: str, port: int = 502, 
                  slave_id: int = 1, product_id: str = None, group_name: str = None, 
                  description: str = None, status: int = 0) -> bool:
        """添加设备信息"""
        try:
            self.cursor.execute('''
                INSERT OR REPLACE INTO devices 
                (device_id, name, ip_address, port, slave_id, product_id, 
                 group_name, description, status, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''', (device_id, name, ip_address, port, slave_id, product_id, 
                  group_name, description, status))
            self.conn.commit()
            logger.info(f"已添加设备: {device_id}")
            return True
        except Exception as e:
            logger.error(f"添加设备失败: {e}")
            self.conn.rollback()
            return False

    def update_device(self, device_id: str, **kwargs) -> bool:
        """更新设备信息"""
        try:
            # 构建更新语句
            set_clauses = []
            params = []
            
            for key, value in kwargs.items():
                if key in ['name', 'ip_address', 'port', 'slave_id', 'product_id', 
                           'group_name', 'description', 'status']:
                    set_clauses.append(f"{key} = ?")
                    params.append(value)
            
            if not set_clauses:
                return True
            
            set_clauses.append("updated_at = CURRENT_TIMESTAMP")
            sql = f"""
                UPDATE devices 
                SET {', '.join(set_clauses)} 
                WHERE device_id = ?
            """
            params.append(device_id)
            
            self.cursor.execute(sql, params)
            self.conn.commit()
            logger.info(f"已更新设备: {device_id}")
            return True
        except Exception as e:
            logger.error(f"更新设备失败: {e}")
            self.conn.rollback()
            return False

    def get_device(self, device_id: str) -> Optional[Dict[str, Any]]:
        """获取设备信息"""
        try:
            self.cursor.execute('''
                SELECT id, device_id, name, ip_address, port, slave_id, product_id, 
                       group_name, description, status, created_at, updated_at
                FROM devices 
                WHERE device_id = ?
            ''', (device_id,))
            
            row = self.cursor.fetchone()
            if not row:
                return None
            
            return {
                'id': row[0],
                'device_id': row[1],
                'name': row[2],
                'ip_address': row[3],
                'port': row[4],
                'slave_id': row[5],
                'product_id': row[6],
                'group_name': row[7],
                'description': row[8],
                'status': row[9],
                'created_at': row[10],
                'updated_at': row[11]
            }
        except Exception as e:
            logger.error(f"获取设备失败: {e}")
            return None

    def get_all_devices(self) -> List[Dict[str, Any]]:
        """获取所有设备信息"""
        try:
            self.cursor.execute('''
                SELECT id, device_id, name, ip_address, port, slave_id, product_id, 
                       group_name, description, status, created_at, updated_at
                FROM devices 
                ORDER BY group_name, name
            ''')
            
            rows = self.cursor.fetchall()
            devices = []
            
            for row in rows:
                devices.append({
                    'id': row[0],
                    'device_id': row[1],
                    'name': row[2],
                    'ip_address': row[3],
                    'port': row[4],
                    'slave_id': row[5],
                    'product_id': row[6],
                    'group_name': row[7],
                    'description': row[8],
                    'status': row[9],
                    'created_at': row[10],
                    'updated_at': row[11]
                })
            
            return devices
        except Exception as e:
            logger.error(f"获取所有设备失败: {e}")
            return []

    def delete_device(self, device_id: str) -> bool:
        """删除设备信息"""
        try:
            self.cursor.execute('''
                DELETE FROM devices 
                WHERE device_id = ?
            ''', (device_id,))
            self.conn.commit()
            logger.info(f"已删除设备: {device_id}")
            return True
        except Exception as e:
            logger.error(f"删除设备失败: {e}")
            self.conn.rollback()
            return False

    def get_system_config(self) -> Dict[str, Any]:
        """获取所有系统配置"""
        try:
            self.cursor.execute('''
                SELECT key, value, description, updated_at
                FROM system_config
                ORDER BY key
            ''')
            rows = self.cursor.fetchall()
            
            config = {}
            for row in rows:
                config[row[0]] = {
                    'value': json.loads(row[1]) if row[1] else None,
                    'description': row[2],
                    'updated_at': row[3]
                }
            
            return config
        except Exception as e:
            logger.error(f"获取系统配置失败: {e}")
            return {}

    def get_system_config_value(self, key: str, default: Any = None) -> Any:
        """获取系统配置值"""
        try:
            self.cursor.execute('''
                SELECT value
                FROM system_config
                WHERE key = ?
            ''', (key,))
            
            row = self.cursor.fetchone()
            if row:
                return json.loads(row[0]) if row[0] else None
            return default
        except Exception as e:
            logger.error(f"获取系统配置值失败: {e}")
            return default

    def set_system_config_value(self, key: str, value: Any, description: Optional[str] = None) -> bool:
        """设置系统配置值"""
        try:
            # 将值转换为JSON字符串
            value_str = json.dumps(value) if value is not None else None
            
            # 检查是否已存在
            self.cursor.execute('''
                SELECT COUNT(*) FROM system_config WHERE key = ?
            ''', (key,))
            
            count = self.cursor.fetchone()[0]
            
            if count > 0:
                # 更新现有配置
                self.cursor.execute('''
                    UPDATE system_config
                    SET value = ?, description = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE key = ?
                ''', (value_str, description, key))
            else:
                # 插入新配置
                self.cursor.execute('''
                    INSERT INTO system_config (key, value, description)
                    VALUES (?, ?, ?)
                ''', (key, value_str, description))
            
            self.conn.commit()
            logger.info(f"已设置系统配置: {key} = {value}")
            return True
        except Exception as e:
            logger.error(f"设置系统配置值失败: {e}")
            self.conn.rollback()
            return False

    def delete_system_config(self, key: str) -> bool:
        """删除系统配置"""
        try:
            self.cursor.execute('''
                DELETE FROM system_config
                WHERE key = ?
            ''', (key,))
            self.conn.commit()
            logger.info(f"已删除系统配置: {key}")
            return True
        except Exception as e:
            logger.error(f"删除系统配置失败: {e}")
            self.conn.rollback()
            return False

    def reset_system_config(self) -> bool:
        """重置系统配置"""
        try:
            self.cursor.execute('''
                DELETE FROM system_config
            ''')
            self.conn.commit()
            logger.info("已重置所有系统配置")
            return True
        except Exception as e:
            logger.error(f"重置系统配置失败: {e}")
            self.conn.rollback()
            return False

    def export_history_data_csv(self, device_id: str, register_address: Optional[int],
                             start_time: Optional[datetime], end_time: Optional[datetime],
                             file_path: str, include_headers: bool = True) -> bool:
        """
        导出历史数据为CSV格式
        
        Args:
            device_id: 设备ID
            register_address: 寄存器地址，None表示所有寄存器
            start_time: 开始时间
            end_time: 结束时间
            file_path: 导出文件路径
            include_headers: 是否包含表头
            
        Returns:
            bool: 导出成功返回True，否则返回False
        """
        try:
            # 获取历史数据
            data = self.get_history_data(device_id, register_address, start_time, end_time, limit=0)
            if not data:
                logger.warning("没有数据可导出")
                return False

            # 确保目录存在
            Path(file_path).parent.mkdir(exist_ok=True)

            # 导出为CSV
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                fieldnames = ['设备ID', '时间戳', '寄存器地址', '寄存器名称', '值', '单位', '状态']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

                if include_headers:
                    writer.writeheader()

                for row in data:
                    writer.writerow({
                        '设备ID': row['device_id'],
                        '时间戳': row['timestamp'],
                        '寄存器地址': row['register_address'],
                        '寄存器名称': row['register_name'],
                        '值': row['value'],
                        '单位': row['unit'] or '',
                        '状态': row['status']
                    })

            logger.info(f"历史数据已导出到CSV: {file_path} ({len(data)}条记录)")
            return True
        except Exception as e:
            logger.error(f"导出CSV失败: {e}")
            return False

    def export_history_data_excel(self, device_id: str, register_address: Optional[int],
                               start_time: Optional[datetime], end_time: Optional[datetime],
                               file_path: str, include_headers: bool = True) -> bool:
        """
        导出历史数据为Excel格式
        
        Args:
            device_id: 设备ID
            register_address: 寄存器地址，None表示所有寄存器
            start_time: 开始时间
            end_time: 结束时间
            file_path: 导出文件路径
            include_headers: 是否包含表头
            
        Returns:
            bool: 导出成功返回True，否则返回False
        """
        if not EXCEL_SUPPORT:
            logger.error("Excel导出功能不可用，缺少openpyxl库")
            return False

        try:
            # 获取历史数据
            data = self.get_history_data(device_id, register_address, start_time, end_time, limit=0)
            if not data:
                logger.warning("没有数据可导出")
                return False

            # 确保目录存在
            Path(file_path).parent.mkdir(exist_ok=True)

            # 创建Excel工作簿
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = '历史数据'

            # 定义表头
            headers = ['设备ID', '时间戳', '寄存器地址', '寄存器名称', '值', '单位', '状态']

            # 设置表头样式
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))

            # 写入表头
            if include_headers:
                for col, header in enumerate(headers, start=1):
                    cell = worksheet.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.fill = openpyxl.styles.PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
                    cell.border = border

            # 写入数据
            for row_idx, row_data in enumerate(data, start=2 if include_headers else 1):
                worksheet.cell(row=row_idx, column=1, value=row_data['device_id'])
                worksheet.cell(row=row_idx, column=2, value=row_data['timestamp'])
                worksheet.cell(row=row_idx, column=3, value=row_data['register_address'])
                worksheet.cell(row=row_idx, column=4, value=row_data['register_name'])
                worksheet.cell(row=row_idx, column=5, value=row_data['value'])
                worksheet.cell(row=row_idx, column=6, value=row_data['unit'] or '')
                worksheet.cell(row=row_idx, column=7, value=row_data['status'])

            # 调整列宽
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                worksheet.column_dimensions[col].auto_size = True

            # 保存工作簿
            workbook.save(file_path)

            logger.info(f"历史数据已导出到Excel: {file_path} ({len(data)}条记录)")
            return True
        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            return False

    def export_all_data(self, file_path: str, format: str = "csv",
                      include_headers: bool = True) -> bool:
        """
        导出所有设备的历史数据
        
        Args:
            file_path: 导出文件路径
            format: 导出格式，'csv'或'excel'
            include_headers: 是否包含表头
            
        Returns:
            bool: 导出成功返回True，否则返回False
        """
        try:
            # 获取所有设备
            devices = self.get_all_devices()
            if not devices:
                logger.warning("没有设备可导出数据")
                return False

            if format == "csv":
                # 为每个设备创建单独的CSV文件
                base_path = Path(file_path)
                for device in devices:
                    device_file_path = base_path.parent / f"{base_path.stem}_{device['device_id']}{base_path.suffix}"
                    self.export_history_data_csv(
                        device_id=device['device_id'],
                        register_address=None,
                        start_time=None,
                        end_time=None,
                        file_path=str(device_file_path),
                        include_headers=include_headers
                    )
            elif format == "excel":
                if not EXCEL_SUPPORT:
                    logger.error("Excel导出功能不可用")
                    return False

                # 创建Excel工作簿
                workbook = openpyxl.Workbook()
                workbook.remove(workbook.active)  # 移除默认工作表

                # 为每个设备创建单独的工作表
                for device in devices:
                    worksheet = workbook.create_sheet(title=device['device_id'][:30])  # 限制工作表名称长度
                    
                    # 获取设备数据
                    data = self.get_history_data(device['device_id'], None, None, None, limit=0)
                    if data:
                        # 写入表头
                        headers = ['时间戳', '寄存器地址', '寄存器名称', '值', '单位', '状态']
                        if include_headers:
                            for col, header in enumerate(headers, start=1):
                                worksheet.cell(row=1, column=col, value=header)
                        
                        # 写入数据
                        for row_idx, row_data in enumerate(data, start=2 if include_headers else 1):
                            worksheet.cell(row=row_idx, column=1, value=row_data['timestamp'])
                            worksheet.cell(row=row_idx, column=2, value=row_data['register_address'])
                            worksheet.cell(row=row_idx, column=3, value=row_data['register_name'])
                            worksheet.cell(row=row_idx, column=4, value=row_data['value'])
                            worksheet.cell(row=row_idx, column=5, value=row_data['unit'] or '')
                            worksheet.cell(row=row_idx, column=6, value=row_data['status'])
                        
                        # 调整列宽
                        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                            worksheet.column_dimensions[col].auto_size = True

                # 保存工作簿
                workbook.save(file_path)
                logger.info(f"所有设备数据已导出到Excel: {file_path}")
            else:
                logger.error(f"不支持的导出格式: {format}")
                return False

            return True
        except Exception as e:
            logger.error(f"导出所有数据失败: {e}")
            return False

    def add_history_data(self, device_id: str, timestamp: datetime, register_address: int, 
                        register_name: str, value: float, unit: Optional[str] = None, 
                        status: int = 0) -> bool:
        """添加历史数据"""
        try:
            self.cursor.execute('''
                INSERT INTO device_history 
                (device_id, timestamp, register_address, register_name, value, unit, status)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (device_id, timestamp, register_address, register_name, value, unit, status))
            self.conn.commit()
            return True
        except Exception as e:
            logger.error(f"添加历史数据失败: {e}")
            self.conn.rollback()
            return False

    def batch_add_history_data(self, data_list: List[Dict[str, Any]]) -> bool:
        """批量添加历史数据"""
        if not data_list:
            return True

        try:
            # 准备批量插入数据
            insert_data = []
            for data in data_list:
                insert_data.append((
                    data['device_id'],
                    data['timestamp'],
                    data['register_address'],
                    data['register_name'],
                    data['value'],
                    data.get('unit'),
                    data.get('status', 0)
                ))

            # 批量插入
            self.cursor.executemany('''
                INSERT INTO device_history 
                (device_id, timestamp, register_address, register_name, value, unit, status)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', insert_data)

            self.conn.commit()
            logger.info(f"批量添加历史数据完成: {len(data_list)} 条记录")
            return True
        except Exception as e:
            logger.error(f"批量添加历史数据失败: {e}")
            self.conn.rollback()
            return False

    def get_history_data(self, device_id: str, register_address: Optional[int] = None, 
                        start_time: Optional[datetime] = None, 
                        end_time: Optional[datetime] = None, 
                        limit: int = 1000) -> List[Dict[str, Any]]:
        """获取历史数据"""
        try:
            # 构建查询语句
            query = '''
                SELECT id, device_id, timestamp, register_address, register_name, value, unit, status
                FROM device_history 
                WHERE device_id = ?
            '''
            params = [device_id]

            if register_address is not None:
                query += " AND register_address = ?"
                params.append(register_address)

            if start_time is not None:
                query += " AND timestamp >= ?"
                params.append(start_time)

            if end_time is not None:
                query += " AND timestamp <= ?"
                params.append(end_time)

            query += " ORDER BY timestamp DESC LIMIT ?"
            params.append(limit)

            # 执行查询
            self.cursor.execute(query, params)
            rows = self.cursor.fetchall()

            # 处理结果
            history_data = []
            for row in rows:
                history_data.append({
                    'id': row[0],
                    'device_id': row[1],
                    'timestamp': row[2],
                    'register_address': row[3],
                    'register_name': row[4],
                    'value': row[5],
                    'unit': row[6],
                    'status': row[7]
                })

            return history_data
        except Exception as e:
            logger.error(f"获取历史数据失败: {e}")
            return []

    def get_latest_data(self, device_id: str) -> Dict[int, Dict[str, Any]]:
        """获取设备最新数据"""
        try:
            self.cursor.execute('''
                SELECT register_address, register_name, value, unit, status, timestamp
                FROM device_history 
                WHERE device_id = ?
                GROUP BY register_address
                HAVING timestamp = MAX(timestamp)
            ''', (device_id,))

            rows = self.cursor.fetchall()
            latest_data = {}

            for row in rows:
                latest_data[row[0]] = {
                    'register_name': row[1],
                    'value': row[2],
                    'unit': row[3],
                    'status': row[4],
                    'timestamp': row[5]
                }

            return latest_data
        except Exception as e:
            logger.error(f"获取最新数据失败: {e}")
            return {}

    def add_alarm(self, device_id: str, alarm_id: str, alarm_name: str, 
                 register_address: int, register_name: str, trigger_value: float, 
                 threshold: float, condition: str, severity: str, message: str, 
                 status: str = 'active') -> bool:
        """添加报警记录"""
        try:
            self.cursor.execute('''
                INSERT INTO alarms 
                (device_id, alarm_id, alarm_name, register_address, register_name, 
                 trigger_value, threshold, condition, severity, message, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (device_id, alarm_id, alarm_name, register_address, register_name, 
                  trigger_value, threshold, condition, severity, message, status))
            self.conn.commit()
            logger.info(f"已添加报警记录: {alarm_id} - {device_id}")
            return True
        except Exception as e:
            logger.error(f"添加报警记录失败: {e}")
            self.conn.rollback()
            return False

    def clear_alarm(self, alarm_id: str, device_id: str) -> bool:
        """清除报警记录"""
        try:
            self.cursor.execute('''
                UPDATE alarms 
                SET status = 'cleared', cleared_at = CURRENT_TIMESTAMP
                WHERE alarm_id = ? AND device_id = ? AND status = 'active'
            ''', (alarm_id, device_id))
            self.conn.commit()
            logger.info(f"已清除报警记录: {alarm_id} - {device_id}")
            return True
        except Exception as e:
            logger.error(f"清除报警记录失败: {e}")
            self.conn.rollback()
            return False

    def get_active_alarms(self, device_id: Optional[str] = None) -> List[Dict[str, Any]]:
        """获取活动报警记录"""
        try:
            query = '''
                SELECT id, device_id, alarm_id, alarm_name, register_address, register_name, 
                       trigger_value, threshold, condition, severity, message, status, timestamp
                FROM alarms 
                WHERE status = 'active'
            '''
            params = []

            if device_id is not None:
                query += " AND device_id = ?"
                params.append(device_id)

            query += " ORDER BY timestamp DESC"

            self.cursor.execute(query, params)
            rows = self.cursor.fetchall()

            alarms = []
            for row in rows:
                alarms.append({
                    'id': row[0],
                    'device_id': row[1],
                    'alarm_id': row[2],
                    'alarm_name': row[3],
                    'register_address': row[4],
                    'register_name': row[5],
                    'trigger_value': row[6],
                    'threshold': row[7],
                    'condition': row[8],
                    'severity': row[9],
                    'message': row[10],
                    'status': row[11],
                    'timestamp': row[12]
                })

            return alarms
        except Exception as e:
            logger.error(f"获取活动报警记录失败: {e}")
            return []

    def add_communication_log(self, device_id: Optional[str], host: str, port: int, 
                             direction: str, function_code: Optional[int] = None, 
                             address: Optional[int] = None, count: Optional[int] = None, 
                             data: Optional[Dict[str, Any]] = None, raw_data: str = None, 
                             success: bool = True, error_message: Optional[str] = None) -> bool:
        """添加通信日志"""
        try:
            # 转换数据为JSON字符串
            data_json = json.dumps(data) if data else None

            self.cursor.execute('''
                INSERT INTO communication_logs 
                (device_id, host, port, direction, function_code, address, count, 
                 data, raw_data, success, error_message)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (device_id, host, port, direction, function_code, address, count, 
                  data_json, raw_data, 1 if success else 0, error_message))
            self.conn.commit()
            return True
        except Exception as e:
            logger.error(f"添加通信日志失败: {e}")
            self.conn.rollback()
            return False

    def get_communication_logs(self, device_id: Optional[str] = None, 
                              start_time: Optional[datetime] = None, 
                              end_time: Optional[datetime] = None, 
                              limit: int = 1000) -> List[Dict[str, Any]]:
        """获取通信日志"""
        try:
            query = '''
                SELECT id, device_id, host, port, direction, function_code, address, count, 
                       data, raw_data, timestamp, success, error_message
                FROM communication_logs 
            '''
            params = []
            where_clauses = []

            if device_id is not None:
                where_clauses.append("device_id = ?")
                params.append(device_id)

            if start_time is not None:
                where_clauses.append("timestamp >= ?")
                params.append(start_time)

            if end_time is not None:
                where_clauses.append("timestamp <= ?")
                params.append(end_time)

            if where_clauses:
                query += " WHERE " + " AND ".join(where_clauses)

            query += " ORDER BY timestamp DESC LIMIT ?"
            params.append(limit)

            self.cursor.execute(query, params)
            rows = self.cursor.fetchall()

            logs = []
            for row in rows:
                # 解析JSON数据
                data = None
                if row[8]:
                    try:
                        data = json.loads(row[8])
                    except:
                        data = row[8]

                logs.append({
                    'id': row[0],
                    'device_id': row[1],
                    'host': row[2],
                    'port': row[3],
                    'direction': row[4],
                    'function_code': row[5],
                    'address': row[6],
                    'count': row[7],
                    'data': data,
                    'raw_data': row[9],
                    'timestamp': row[10],
                    'success': bool(row[11]),
                    'error_message': row[12]
                })

            return logs
        except Exception as e:
            logger.error(f"获取通信日志失败: {e}")
            return []

    def get_system_config(self, key: str) -> Optional[str]:
        """获取系统配置"""
        try:
            self.cursor.execute('''
                SELECT value FROM system_config WHERE key = ?
            ''', (key,))
            result = self.cursor.fetchone()
            return result[0] if result else None
        except Exception as e:
            logger.error(f"获取系统配置失败: {e}")
            return None

    def set_system_config(self, key: str, value: str, description: Optional[str] = None) -> bool:
        """设置系统配置"""
        try:
            self.cursor.execute('''
                INSERT OR REPLACE INTO system_config 
                (key, value, description, updated_at)
                VALUES (?, ?, ?, CURRENT_TIMESTAMP)
            ''', (key, value, description))
            self.conn.commit()
            logger.info(f"已设置系统配置: {key} = {value}")
            return True
        except Exception as e:
            logger.error(f"设置系统配置失败: {e}")
            self.conn.rollback()
            return False

    def clean_old_data(self, days: int = 30) -> int:
        """清理旧数据"""
        try:
            # 计算清理时间点
            cutoff_time = datetime.now() - timedelta(days=days)

            # 清理历史数据
            self.cursor.execute('''
                DELETE FROM device_history WHERE timestamp < ?
            ''', (cutoff_time,))
            history_rows_deleted = self.cursor.rowcount

            # 清理通信日志
            self.cursor.execute('''
                DELETE FROM communication_logs WHERE timestamp < ?
            ''', (cutoff_time,))
            log_rows_deleted = self.cursor.rowcount

            # 清理已清除的报警记录
            self.cursor.execute('''
                DELETE FROM alarms WHERE status = 'cleared' AND cleared_at < ?
            ''', (cutoff_time,))
            alarm_rows_deleted = self.cursor.rowcount

            self.conn.commit()

            total_deleted = history_rows_deleted + log_rows_deleted + alarm_rows_deleted
            logger.info(f"清理旧数据完成: 历史数据({history_rows_deleted}), \
                       通信日志({log_rows_deleted}), 报警记录({alarm_rows_deleted})")
            return total_deleted

        except Exception as e:
            logger.error(f"清理旧数据失败: {e}")
            self.conn.rollback()
            return 0

    def export_data(self, table_name: str, output_path: str, 
                   start_time: Optional[datetime] = None, 
                   end_time: Optional[datetime] = None) -> bool:
        """导出数据到CSV文件"""
        import csv

        try:
            # 构建查询
            if table_name == "history":
                query = "SELECT * FROM device_history"
                params = []
                if start_time:
                    query += " WHERE timestamp >= ?"
                    params.append(start_time)
                if end_time:
                    if start_time:
                        query += " AND timestamp <= ?"
                    else:
                        query += " WHERE timestamp <= ?"
                    params.append(end_time)
                query += " ORDER BY timestamp"
            elif table_name == "alarms":
                query = "SELECT * FROM alarms ORDER BY timestamp"
                params = []
            elif table_name == "logs":
                query = "SELECT * FROM communication_logs"
                params = []
                if start_time:
                    query += " WHERE timestamp >= ?"
                    params.append(start_time)
                if end_time:
                    if start_time:
                        query += " AND timestamp <= ?"
                    else:
                        query += " WHERE timestamp <= ?"
                    params.append(end_time)
                query += " ORDER BY timestamp"
            else:
                logger.error(f"不支持的导出表: {table_name}")
                return False

            # 执行查询
            self.cursor.execute(query, params)
            rows = self.cursor.fetchall()

            # 获取列名
            column_names = [description[0] for description in self.cursor.description]

            # 导出到CSV
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(column_names)
                writer.writerows(rows)

            logger.info(f"数据导出完成: {output_path} ({len(rows)} 条记录)")
            return True

        except Exception as e:
            logger.error(f"数据导出失败: {e}")
            return False


# 创建数据库管理器实例
def create_database_manager(db_path: str = "data/equipment_management.db") -> DatabaseManager:
    """创建数据库管理器实例"""
    return DatabaseManager(db_path)


# 测试数据库功能
if __name__ == "__main__":
    try:
        db = DatabaseManager("test.db")
        
        # 测试添加设备
        db.add_device(
            device_id="Pump-01",
            name="Pump-01",
            ip_address="192.168.1.101",
            port=502,
            slave_id=1,
            product_id="pump_station_a",
            group_name="泵站A区",
            description="主泵机组"
        )
        
        # 测试添加历史数据
        now = datetime.now()
        db.batch_add_history_data([
            {
                "device_id": "Pump-01",
                "timestamp": now,
                "register_address": 1,
                "register_name": "温度传感器",
                "value": 25.5,
                "unit": "°C",
                "status": 0
            },
            {
                "device_id": "Pump-01",
                "timestamp": now,
                "register_address": 2,
                "register_name": "压力变送器",
                "value": 1.23,
                "unit": "MPa",
                "status": 0
            }
        ])
        
        # 测试查询数据
        devices = db.get_all_devices()
        print(f"设备列表: {devices}")
        
        history_data = db.get_history_data("Pump-01", limit=10)
        print(f"历史数据: {history_data}")
        
        # 清理测试数据库
        db.close()
        os.remove("test.db")
        
    except Exception as e:
        print(f"测试失败: {e}")
