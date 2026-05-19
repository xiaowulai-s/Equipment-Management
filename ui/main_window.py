# -*- coding: utf-8 -*-
"""
Main Window - Industrial Equipment Management System
Refactored with maintainable constants and clear text mappings
"""

from __future__ import annotations

import re
import webbrowser
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any, Dict, List, Optional

from PySide6.QtCore import QDateTime, QSettings, QSize, Qt, QTimer, Slot, Signal, QObject, QRunnable, QThreadPool
from PySide6.QtGui import QFont, QIcon
from PySide6.QtWidgets import QApplication, QTableWidget  # noqa: F401  (base class of DataTable)
from PySide6.QtWidgets import (
    QAbstractItemView,
    QDateTimeEdit,
    QDialog,
    QFileDialog,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMainWindow,
    QMenu,
    QMessageBox,
    QPlainTextEdit,
    QProgressBar,
    QProgressDialog,
    QPushButton,
    QSizePolicy,
    QSplitter,
    QStackedWidget,
    QTableWidgetItem,
    QTabWidget,
    QTextEdit,
    QToolBar,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
)

from core.data import DatabaseManager
from core.device.device_manager import DeviceManager
from core.device.device_model import DeviceStatus
from core.utils.logger import get_logger
from core.utils.qt_helpers import is_valid_qt_object, safe_reconnect
from core.utils.write_operation_manager import WriteOperationManager
from core.version import __version__
from ui.add_device_dialog import AddDeviceDialog
from ui.app_styles import AppStyles
from ui.batch_operations_dialog import BatchOperationsDialog
from ui.core import ThemeManager
from ui.device_type_dialogs import DeviceTypeDialog
from ui.register_write_dialog import RegisterWriteDialog
from ui.widgets.dynamic_monitor_panel import DynamicMonitorPanel
from ui.dialogs.mcgs_config_dialog import MCGSConfigDialog
from ui.dialogs.device_scan_dialog import DeviceScanDialog
from ui.controllers.device_controller import DeviceController
from ui.widgets import (
    DangerButton,
    DataCard,
    DataTable,
    DeviceTree,
    GhostButton,
    LineEdit,
    PrimaryButton,
    SecondaryButton,
    StatusBadge,
    SuccessButton,
)


if TYPE_CHECKING:
    from core.device.device_model import Device  # noqa: F401

logger = get_logger("main_window")


class TextConstants:
    """UI text constants - centralized for easy maintenance"""

    APP_NAME = "MCGS_EMS"
    APP_VERSION = __version__
    WINDOW_TITLE = f"{APP_NAME} v{APP_VERSION}"

    DEVICE_LIST_TITLE = "设备列表"
    ADD_DEVICE_BTN = "添加设备"
    REMOVE_DEVICE_BTN = "移除设备"
    SEARCH_PLACEHOLDER = "搜索设备名称..."

    WELCOME_TITLE = "欢迎使用MCGS_EMS"
    WELCOME_SUBTITLE = "请从左侧面板选择设备进行监控"

    DEVICE_MONITOR_TITLE = "设备监控"
    DEVICE_NAME_LABEL = "设备名称:"
    STATUS_LABEL = "状态:"
    LAST_UPDATE_LABEL = "最后更新:"

    TAB_REALTIME_DATA = "实时数据"
    TAB_REGISTERS = "寄存器"

    TREE_HEADER_NAME = "设备名称"
    TREE_HEADER_ID = "设备 ID"
    TREE_HEADER_STATUS = "状态"
    TREE_HEADER_ACTIONS = "操作"

    BTN_EDIT = "编辑"
    BTN_CONNECT = "连接"
    BTN_DISCONNECT = "断开"

    MENU_FILE = "文件 (&F)"
    MENU_TOOLS = "工具 (&T)"
    MENU_HELP = "帮助 (&H)"

    ACTION_DEVICE_TYPE_MGMT = "设备类型管理 (&T)"
    ACTION_DATA_EXPORT = "数据导出 (&E)"
    ACTION_EXPORT_DEVICE_CONFIG = "设备配置导出 (&O)"
    ACTION_IMPORT_DEVICE_CONFIG = "设备配置导入 (&I)"
    ACTION_BATCH_OPS = "批量操作"
    ACTION_EXIT = "退出 (&X)"
    ACTION_ABOUT = "关于 (&A)"


class StatusText:
    """Device status text and badge type mappings"""

    STATUS_MAP = {
        DeviceStatus.CONNECTED: ("已连接", "success"),
        DeviceStatus.DISCONNECTED: ("未连接", "warning"),
        DeviceStatus.CONNECTING: ("连接中...", "info"),
        DeviceStatus.ERROR: ("错误", "error"),
    }

    @classmethod
    def get_text(cls, status: int) -> str:
        return cls.STATUS_MAP.get(status, ("Unknown", "default"))[0]

    @classmethod
    def get_badge_type(cls, status: int) -> str:
        return cls.STATUS_MAP.get(status, ("Unknown", "default"))[1]

    @classmethod
    def get_text_with_badge(cls, status: int) -> tuple:
        return cls.STATUS_MAP.get(status, ("Unknown", "default"))


class LogMessages:
    """Log message templates"""

    APP_STARTUP = "应用程序已启动"
    APP_SHUTDOWN = "应用程序已关闭"
    DB_INIT_SUCCESS = "数据库初始化成功"

    DEVICE_ADDED = "设备已添加：{device_id}"
    DEVICE_REMOVED = "设备已移除：{device_id}"
    DEVICE_CONNECTED = "设备已连接：{device_id}"
    DEVICE_DISCONNECTED = "设备已断开：{device_id}"
    DEVICE_CONNECTING = "正在连接设备：{device_id}"
    DEVICE_CONNECT_FAILED = "设备连接失败：{device_id}"

    DEVICE_EDIT_SUCCESS = "设备已更新：{device_id}"
    DEVICE_EDIT_FAILED = "设备更新失败：{device_id}"

    BATCH_OPS_COMPLETE = "批量操作完成：{success}/{total} 成功"

    DATA_EXPORT_SUCCESS = "数据已导出至：{path}"
    DATA_EXPORT_FAILED = "数据导出失败"


class UIMessages:
    """User interface message templates"""

    CONFIRM_DELETE_TITLE = "确认删除"
    CONFIRM_DELETE_MSG = "确定要删除此设备吗？"

    CONNECT_FAILED_TITLE = "连接失败"
    CONNECT_FAILED_MSG = "无法连接到设备"

    SELECT_DEVICE_TITLE = "选择设备"
    SELECT_DEVICE_MSG = "请先选择一个设备"

    EXPORT_SUCCESS_TITLE = "导出成功"
    EXPORT_SUCCESS_MSG = "数据已导出至：{path}"
    DEVICE_CONFIG_EXPORT_SUCCESS_MSG = "设备配置已成功导出至：{path}"

    EXPORT_FAILED_TITLE = "导出失败"
    EXPORT_FAILED_MSG = "数据导出失败，请检查文件格式或权限"
    DEVICE_CONFIG_EXPORT_FAILED_MSG = "设备配置导出失败"

    IMPORT_SUCCESS_TITLE = "导入成功"
    DEVICE_CONFIG_IMPORT_SUCCESS_MSG = "设备配置已成功导入"

    IMPORT_FAILED_TITLE = "导入失败"
    DEVICE_CONFIG_IMPORT_FAILED_MSG = "设备配置导入失败"

    IMPORT_CONFIRM_TITLE = "确认导入"
    DEVICE_CONFIG_IMPORT_CONFIRM_MSG = "确定要导入设备配置吗？\n现有设备将被覆盖。"

    GITHUB_REPO_URL = "https://github.com/xiaowulai-s/Equipment-Management"

    ABOUT_TITLE = "关于"
    ABOUT_MSG = (
        f"{TextConstants.APP_NAME} v{TextConstants.APP_VERSION}<br><br>"
        "基于 PySide6 和 Modbus TCP 协议的工业设备监控上位机软件<br>"
        "四层解耦架构: UI → Controller → Service → Core<br><br>"
        "功能特性:<br>"
        "━━━━━━━━━━━━━━━━━━━━━<br>"
        "• MCGS 触摸屏 Modbus TCP 通信<br>"
        "• 多设备并发轮询与实时监控<br>"
        "• DataCard 数据卡片 | 实时曲线图<br>"
        "• 寄存器表格 | 数据日志 | 历史查询<br>"
        "• 设备增删改查 | 批量管理<br>"
        "• 报警检测 | 数据导出 (CSV/Excel/JSON)<br>"
        "• 浅色/深色主题切换<br>"
        "━━━━━━━━━━━━━━━━━━━━━<br>"
        "技术栈:<br>"
        "  Python 3.10+ | PySide6 | SQLAlchemy<br>"
        "  pymodbus | structlog | QThreadPool<br><br>"
        f"© 2026 QYH MCGS_EMS Team<br>"
        f'代码仓库: <a href="{GITHUB_REPO_URL}" style="color: #3B82F6;">{GITHUB_REPO_URL}</a>'
    )


def _downsample_data(points: List[Tuple[datetime, float]], max_points: int = 5000) -> List[Tuple[datetime, float]]:
    """等间隔降采样，减少图表渲染压力

    Args:
        points: 原始数据点列表 [(timestamp, value), ...]
        max_points: 最多保留的数据点数

    Returns:
        降采样后的数据点列表
    """
    if len(points) <= max_points:
        return points

    step = len(points) / max_points
    result: List[Tuple[datetime, float]] = []
    for i in range(max_points):
        idx = min(int(i * step), len(points) - 1)
        result.append(points[idx])
    return result


class _HistoryQueryTask(QRunnable):
    """历史数据查询任务（在 QThreadPool 中执行）

    负责在后台线程中查询数据库，通过信号回传结果到主线程。
    """

    class Signals(QObject):
        """信号集合"""

        result = Signal(object)  # data_by_param, param_names, has_data
        progress = Signal(object)  # (current_index, total)
        error = Signal(str)  # error_message

    def __init__(
        self, history_service, device_id: str, param_names: List[str], hours: int = 24, start_time=None, end_time=None
    ):
        super().__init__()
        self._hs = history_service
        self._device_id = device_id
        self._param_names = param_names
        self._hours = hours
        self._start_time = start_time
        self._end_time = end_time
        self._cancelled = False
        self.signals = self.Signals()
        self.setAutoDelete(True)

    def cancel(self):
        """取消任务"""
        self._cancelled = True

    def run(self):
        """在后台线程中执行数据库查询"""
        from datetime import datetime as _dt
        import logging as _logging

        _log = _logging.getLogger(__name__)

        try:
            data_by_param: Dict[str, List[Tuple[str, float]]] = {}
            has_data = False
            total = len(self._param_names)

            for idx, pname in enumerate(self._param_names):
                # 检查是否被取消
                if self._cancelled:
                    _log.info("历史查询任务被取消 [dev=%s]", self._device_id)
                    return

                # 发送进度更新
                self.signals.progress.emit((idx + 1, total))

                # 查询数据库
                try:
                    if self._start_time is not None and self._end_time is not None:
                        data = self._hs.query_trend(
                            self._device_id, pname, start_time=self._start_time, end_time=self._end_time
                        )
                    else:
                        data = self._hs.query_trend(self._device_id, pname, hours=self._hours)
                except Exception as e:
                    _log.error("查询历史数据异常 [dev=%s, param=%s]: %s", self._device_id, pname, e)
                    data = None

                if data:
                    has_data = True
                    points_list: List[Tuple[str, float]] = []
                    for timestamp, value in data:
                        if isinstance(timestamp, str):
                            try:
                                dt = _dt.fromisoformat(timestamp)
                            except ValueError:
                                dt = _dt.now()
                        else:
                            dt = timestamp
                        ts_str = dt.strftime("%Y-%m-%d %H:%M:%S")
                        points_list.append((ts_str, float(value)))
                    data_by_param[pname] = points_list

            # 发送最终结果
            self.signals.result.emit((data_by_param, self._param_names, has_data))

        except Exception as e:
            _log.error("后台加载历史数据失败: %s", e, exc_info=True)
            self.signals.error.emit(str(e))


class MainWindow(QMainWindow):
    """
    主窗口 - MCGS_EMS

    提供完整的设备管理功能，包括：
    - 设备列表管理
    - 实时数据监控
    - 报警系统
    - 数据导出
    - 批量操作
    - 主题切换

    Attributes:
        _db_manager: 数据库管理器
        _device_manager: 设备管理器
        _theme_manager: 主题管理器
        _current_device_id: 当前选中设备ID
    """

    def __init__(self, db_manager: Optional[DatabaseManager] = None, parent: Optional[QWidget] = None) -> None:
        """
        初始化主窗口

        Args:
            db_manager: 数据库管理器实例，为None时自动创建
            parent: 父窗口
        """
        super().__init__(parent)

        self._set_app_icon()
        self._db_manager = db_manager or DatabaseManager()
        self._device_manager = DeviceManager(db_manager=self._db_manager)
        self._device_controller = DeviceController(self._device_manager, parent=self)
        self._theme_manager = ThemeManager()
        self._sort_order = Qt.SortOrder.AscendingOrder
        self._sort_column = 0
        self._current_device_id: Optional[str] = None
        self._group_by_type: bool = False  # 设备树是否按类型分组
        self._left_panel_collapsed = False
        self._left_panel_saved_size = 520  # 从480增加到520，确保设备列表有足够空间显示

        # 操作防重复点击锁（P0-2修复）
        self._operation_lock = False
        self._current_operation_btn = None  # 记录当前操作的按钮，用于恢复状态

        # 数据卡片配置存储: {device_id: [card_configs]}
        self._device_cards: Dict[str, List[Dict[str, Any]]] = {}

        # 趋势图配置存储: {device_id: [chart_configs]}
        self._device_charts: Dict[str, List[Dict[str, Any]]] = {}

        # v3.2 新增：动态监控面板存储 {device_id: DynamicMonitorPanel实例}
        self._monitor_panels: Dict[str, DynamicMonitorPanel] = {}

        # v3.2 新增：写操作管理器（安全网关）
        self._write_manager = WriteOperationManager(self)
        self._pending_write_req_id: Optional[str] = None  # 当前待确认的写请求ID

        # ✅ 新增：权限管理器（三级权限体系）
        from core.utils.permission_manager import PermissionManager

        self._permission_mgr = PermissionManager(parent=self)
        self._permission_mgr.permission_changed.connect(self._on_permission_changed)
        self._permission_mgr.session_timeout.connect(self._on_session_timeout)

        # 将权限管理器注入到写操作管理器
        self._write_manager.set_permission_manager(self._permission_mgr)

        # ✅ 新增：操作撤销管理器
        from core.utils.operation_undo_manager import OperationUndoManager

        self._undo_mgr = OperationUndoManager(max_history=100, parent=self)
        self._undo_req_map: Dict[str, str] = {}
        self._undo_mgr.undo_executed.connect(self._on_undo_executed)
        self._undo_mgr.undo_failed.connect(self._on_undo_failed)

        # ✅ 新增：将撤销管理器注入到写操作管理器
        self._write_manager.set_undo_manager(self._undo_mgr)

        # ✅ 新增：MCGS快速连接模块 — 通过 MCGSController 异步调度
        self._init_mcgsm_controller()

        # 数据清理调度器
        from core.data.cleanup_scheduler import CleanupScheduler

        self._cleanup_scheduler = CleanupScheduler(self._db_manager)
        self._cleanup_scheduler.start()

        # UI 偏好持久化服务
        from core.services.ui_preferences_service import UIPreferencesService

        self._ui_prefs = UIPreferencesService()

        from ui.controllers.status_bar_controller import StatusBarController
        from ui.controllers.monitor_page_controller import MonitorPageController

        self._status_bar_controller = StatusBarController()
        self._monitor_controller = MonitorPageController()

        self._init_ui()
        self._connect_signals()
        self._refresh_device_list()
        self._apply_theme()
        self._load_ui_preferences()
        self._cleaned = False

        logger.info(LogMessages.APP_STARTUP)

    def __del__(self) -> None:
        if getattr(self, "_cleaned", True):
            return
        self._cleaned = True

        try:
            if hasattr(self, "_cleanup_scheduler"):
                self._cleanup_scheduler.stop()
                del self._cleanup_scheduler
        except Exception:
            pass

        try:
            if hasattr(self, "_device_manager"):
                for device_info in list(self._device_manager.get_all_devices()):
                    try:
                        device_id = device_info.get("device_id")
                        if device_id:
                            self._device_manager.disconnect_device(device_id)
                    except Exception as e:
                        if "Signal source has been deleted" not in str(e):
                            logger.error("断开设备连接失败: %s", str(e))
        except Exception:
            pass

        try:
            if hasattr(self, "_db_manager"):
                del self._db_manager
        except Exception:
            pass

        try:
            if hasattr(self, "_device_cards"):
                self._device_cards.clear()
            if hasattr(self, "_device_charts"):
                self._device_charts.clear()
        except Exception:
            pass

    def _init_ui(self) -> None:
        self.setWindowTitle(TextConstants.WINDOW_TITLE)
        # 设置合理的最小窗口尺寸（支持自适应调整）
        self.setMinimumSize(1024, 768)

        # 获取屏幕可用区域，智能设置初始窗口大小
        from PySide6.QtWidgets import QApplication

        screen = QApplication.primaryScreen().availableGeometry()
        screen_width = screen.width()
        screen_height = screen.height()

        # 初始窗口大小：屏幕的80%，但不超过1600x900
        init_width = min(int(screen_width * 0.8), 1600)
        init_height = min(int(screen_height * 0.8), 900)
        self.resize(init_width, init_height)

        self._settings = QSettings("MCGS_EMS", "EquipmentManagement")
        geom = self._settings.value("window/geometry")
        if geom is not None:
            self.restoreGeometry(geom)

        self._create_menu_bar()

        self._central_widget = QWidget()
        self.setCentralWidget(self._central_widget)

        main_layout = QVBoxLayout(self._central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        self._create_tool_bar()

        # ── QSplitter: 左面板 + 右面板 ──
        self._splitter = QSplitter(Qt.Orientation.Horizontal)
        main_layout.addWidget(self._splitter)

        # 左侧面板
        self._left_panel = self._create_left_panel()
        self._splitter.addWidget(self._left_panel)

        # 中间面板 (StackedWidget: 页面容器)
        self._stack_widget = QStackedWidget()
        self._stack_widget.setStyleSheet(AppStyles.STACKED_WIDGET)
        self._stack_widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        # 页面0: 欢迎页
        welcome_page = self._create_welcome_page()
        self._stack_widget.addWidget(welcome_page)

        # 页面1: 设备监控页（内嵌命令终端）
        monitor_page = self._create_monitor_page()
        self._stack_widget.addWidget(monitor_page)

        self._splitter.addWidget(self._stack_widget)

        # 默认显示欢迎页 (index 0)
        self._stack_widget.setCurrentIndex(0)

        # Splitter 配置（支持自适应调整，增加左侧面板权重）
        self._splitter.setStretchFactor(0, 30)  # 左侧面板占30%
        self._splitter.setStretchFactor(1, 70)  # 中间面板占70%

        # 初始尺寸比例：左30% + 中70%
        total_width = max(self.width(), 1024)
        initial_sizes = [
            int(total_width * 0.30),
            int(total_width * 0.70),
        ]
        self._splitter.setSizes(initial_sizes)

        # 允许面板折叠
        self._splitter.setCollapsible(0, True)  # 左侧面板可折叠
        self._splitter.setCollapsible(1, False)  # 中间主区域不可折叠

        # 设置分割器的最小尺寸约束（防止面板被压缩到不可用）
        self._splitter.setHandleWidth(3)  # 细分割线，更现代的外观

        self._splitter.setStyleSheet(AppStyles.SPLITTER)
        self._splitter.splitterMoved.connect(self._on_splitter_moved)

        # ── 折叠/展开浮动按钮 ──
        self._create_collapse_buttons()

        self._create_status_bar()

    def _create_menu_bar(self) -> None:
        menubar = self.menuBar()

        file_menu = menubar.addMenu(TextConstants.MENU_FILE)

        device_type_action = file_menu.addAction(TextConstants.ACTION_DEVICE_TYPE_MGMT)
        device_type_action.triggered.connect(self._show_device_type_dialog)

        file_menu.addSeparator()

        # 设备配置导出/导入选项
        export_config_action = file_menu.addAction(TextConstants.ACTION_EXPORT_DEVICE_CONFIG)
        export_config_action.triggered.connect(self._show_export_device_config_dialog)

        import_config_action = file_menu.addAction(TextConstants.ACTION_IMPORT_DEVICE_CONFIG)
        import_config_action.triggered.connect(self._show_import_device_config_dialog)

        file_menu.addSeparator()

        history_action = file_menu.addAction("历史数据 (&H)")
        history_action.triggered.connect(self._on_mcgsm_show_history)

        file_menu.addSeparator()

        exit_action = file_menu.addAction(TextConstants.ACTION_EXIT)
        exit_action.triggered.connect(self.close)

        help_menu = menubar.addMenu(TextConstants.MENU_HELP)

        about_action = help_menu.addAction(TextConstants.ACTION_ABOUT)
        about_action.triggered.connect(self._show_about)

    def _create_tool_bar(self) -> None:
        toolbar = QToolBar("Main Toolbar")
        toolbar.setMovable(False)
        toolbar.setStyleSheet(AppStyles.TOOLBAR)
        self.addToolBar(toolbar)

    def _create_status_bar(self) -> None:
        self._status_bar_controller.build(self, AppStyles.__dict__)

        self._status_msg_label = self._status_bar_controller._status_msg_label
        self._status_total_label = self._status_bar_controller._status_total_label
        self._status_online_label = self._status_bar_controller._status_online_label
        self._status_offline_label = self._status_bar_controller._status_offline_label
        self._status_error_label = self._status_bar_controller._status_error_label

        # 最后更新时间
        self._status_time_label = self._make_status_label("更新时间: 00:00:00", "#6B7280")
        self.statusBar().addPermanentWidget(self._status_time_label)

    @staticmethod
    def _make_status_label(text: str, color: str) -> QLabel:
        """创建状态栏统计标签"""
        label = QLabel(text)
        label.setStyleSheet(f"color: {color}; font-size: 12px; font-weight: 500; padding: 0 8px;")
        return label

    # ═══════════════════════════════════════════════════════════
    # 折叠/展开按钮系统
    # ═══════════════════════════════════════════════════════════

    def _create_collapse_buttons(self) -> None:
        """初始化折叠和展开按钮的样式和状态.

        按钮已经在左侧面板和监控页面中创建，这里只处理样式和初始状态。
        """
        # 应用按钮样式
        if hasattr(self, "_collapse_btn"):
            self._apply_edge_btn_style(self._collapse_btn)

        if hasattr(self, "_expand_btn"):
            self._apply_edge_btn_style(self._expand_btn)
            self._expand_btn.hide()  # 默认隐藏，面板折叠后显示

    def _apply_edge_btn_style(self, btn: QPushButton) -> None:
        """为边缘按钮设置主题感知的圆角样式."""
        colors = self._theme_manager.colors

        # 根据按钮对象名称判断类型
        btn_name = btn.objectName()

        if btn_name in ["left_collapse_btn", "left_expand_btn"]:
            # 标题栏左侧按钮：圆形，无边框，悬停效果
            btn.setStyleSheet(
                f"""
                QPushButton {{
                    background: transparent;
                    color: {colors.text_secondary};
                    border: none;
                    border-radius: 50%;
                    padding: 4px;
                    font-size: 12px;
                    text-align: center;
                }}
                QPushButton:hover {{
                    background: {colors.bg_hover};
                    color: {colors.text_primary};
                }}
            """
            )
        elif btn_name.startswith("right_"):
            # 右侧边沿按钮：左侧圆角
            bg = colors.bg_hover
            border_top = f"1px solid {colors.border_default}"
            border_left = f"1px solid {colors.border_default}"
            border_bottom = f"1px solid {colors.border_default}"
            btn.setStyleSheet(
                f"""
                QPushButton {{
                    background: {bg};
                    color: {colors.text_secondary};
                    border-top: {border_top};
                    border-left: {border_left};
                    border-bottom: {border_bottom};
                    border-right: none;
                    border-top-left-radius: 6px;
                    border-bottom-left-radius: 6px;
                    padding: 0px 4px;
                    font-size: 10px;
                }}
                QPushButton:hover {{
                    background: {colors.bg_overlay};
                    color: {colors.text_primary};
                }}
            """
            )

    def _reposition_edge_buttons(self) -> None:
        """重新定位折叠/展开按钮."""
        # 移除了对报文工具浮动按钮的处理，因为现在使用标题栏按钮

    def _collapse_left_panel(self) -> None:
        """折叠左侧面板."""
        current_sizes = self._splitter.sizes()
        self._left_panel_saved_size = current_sizes[0]
        total = self._splitter.width()
        self._splitter.setSizes([0, total])
        self._left_panel.setVisible(False)
        self._collapse_btn.hide()
        self._expand_btn.show()
        self._left_panel_collapsed = True
        QTimer.singleShot(50, self._reposition_edge_buttons)

        QTimer.singleShot(50, lambda: self._adjust_vertical_splitter())

    def _expand_left_panel(self) -> None:
        """展开左侧面板（自适应当前窗口大小）."""
        self._left_panel.setVisible(True)
        self._expand_btn.hide()
        self._collapse_btn.show()
        self._left_panel_collapsed = False

        total = self._splitter.width()

        # 智能恢复宽度：根据当前窗口大小动态计算（确保设备列表完整显示）
        if hasattr(self, "_left_panel_saved_size") and self._left_panel_saved_size > 0:
            # 使用保存的宽度，但不超过当前窗口的合理比例
            saved = self._left_panel_saved_size
            max_left = int(total * 0.35)  # 左侧最多占35%（从30%增加）
            left_size = min(saved, max_left)
            # 确保最小宽度足够显示设备树
            left_size = max(left_size, int(total * 0.25))  # 至少占25%
        else:
            # 无保存记录时使用默认比例（优化后的值）
            if total < 1280:
                left_size = int(total * 0.25)  # 小屏幕25%（从18%增加）
            elif total < 1600:
                left_size = int(total * 0.28)  # 中等屏幕28%（从20%增加）
            else:
                left_size = int(total * 0.30)  # 大屏幕30%（从22%增加）

        min_middle = int(total * 0.45)
        middle_width = total - left_size
        if middle_width < min_middle:
            left_size = total - min_middle

        self._splitter.setSizes([left_size, total - left_size])

        QTimer.singleShot(50, self._reposition_edge_buttons)
        QTimer.singleShot(50, self._update_tree_adaptive_sizes)

        # 重新调整垂直分割器大小，确保高度比例一致
        QTimer.singleShot(50, lambda: self._adjust_vertical_splitter())

    def _adjust_vertical_splitter(self) -> None:
        """智能调整垂直分割器（监控区+终端）的高度比例."""
        if not hasattr(self, "_right_splitter"):
            return

        vertical_height = self._right_splitter.height()

        # 根据窗口总高度动态调整比例（线性过渡）
        total_height = self.height()

        # 线性插值计算监控区比例: 600px->0.80, 1200px->0.65
        if total_height <= 600:
            monitor_ratio = 0.80  # 小窗口：监控区占更多空间
        elif total_height >= 1200:
            monitor_ratio = 0.65  # 大窗口：给终端更多空间
        else:
            # 线性插值
            monitor_ratio = 0.80 - (total_height - 600) / 600 * 0.15

        monitor_height = int(vertical_height * monitor_ratio)
        log_height = vertical_height - monitor_height

        # 确保最小高度（防止组件被压缩到不可用）
        min_monitor = 300  # 监控区最小高度
        min_terminal = 150  # 终端最小高度

        if monitor_height < min_monitor:
            monitor_height = min_monitor
            log_height = vertical_height - monitor_height
            if log_height < min_terminal:
                log_height = min_terminal

        self._right_splitter.setSizes([monitor_height, log_height])

    def _on_splitter_moved(self, pos: int, index: int) -> None:
        """Splitter 拉动时重新定位按钮并自适应调整（含最小尺寸保护）."""
        if not self._left_panel_collapsed:
            self._reposition_edge_buttons()
            self._update_tree_adaptive_sizes()

        # 重新调整垂直分割器大小，确保高度比例一致
        self._adjust_vertical_splitter()

        # 最小尺寸保护：检查并纠正不合理的面板尺寸
        self._enforce_minimum_panel_sizes()

    def _enforce_minimum_panel_sizes(self) -> None:
        if not hasattr(self, "_splitter"):
            return

        sizes = list(self._splitter.sizes())
        total = sum(sizes)

        if total == 0:
            return

        MIN_LEFT = 300
        MIN_MIDDLE = 500

        adjusted = False

        if sizes[0] > 0 and sizes[0] < MIN_LEFT:
            sizes[0] = MIN_LEFT
            adjusted = True

        if sizes[1] < MIN_MIDDLE:
            deficit = MIN_MIDDLE - sizes[1]
            if sizes[0] > MIN_LEFT:
                borrow = min(deficit, sizes[0] - MIN_LEFT)
                sizes[0] -= borrow
                deficit -= borrow
            sizes[1] = MIN_MIDDLE
            adjusted = True

        if adjusted:
            diff = total - sum(sizes)
            if diff != 0:
                sizes[1] += diff
            self._splitter.setSizes(sizes)

    def resizeEvent(self, event) -> None:
        """窗口缩放时智能调整布局 - 自适应响应式设计."""
        super().resizeEvent(event)
        # 使用 QTimer.singleShot 确保在布局完成后执行调整，避免闪烁
        QTimer.singleShot(0, self._reposition_edge_buttons)
        QTimer.singleShot(0, self._adjust_vertical_splitter)
        # 新增：智能自适应布局调整
        QTimer.singleShot(0, self._adaptive_layout_adjustment)

    def _adaptive_layout_adjustment(self) -> None:
        """根据窗口尺寸智能调整各组件布局比例."""
        if not hasattr(self, "_splitter"):
            return

        current_width = self.width()

        # 获取当前分割器状态
        sizes = list(self._splitter.sizes())
        total = sum(sizes)

        if total == 0:
            return

        # 自适应策略：根据窗口宽度动态调整面板比例（线性过渡）
        # 线性插值: 800px->0.25, 1600px->0.35
        if current_width <= 800:
            left_ratio = 0.25  # 小窗口：左侧最小比例
        elif current_width >= 1600:
            left_ratio = 0.35  # 大窗口：左侧更宽松
        else:
            # 线性插值
            left_ratio = 0.25 + (current_width - 800) / 800 * 0.10

        left_width = int(total * left_ratio)
        middle_width = total - left_width
        self._splitter.setSizes([left_width, middle_width])

        # 动态调整左侧面板最小宽度（确保设备树列能完整显示）
        if hasattr(self, "_left_panel"):
            # 线性插值: 800px->300px, 1600px->380px
            if current_width <= 800:
                min_width = 300
            elif current_width >= 1600:
                min_width = 380
            else:
                min_width = int(300 + (current_width - 800) / 800 * 80)
            self._left_panel.setMinimumWidth(min_width)

    def showEvent(self, event) -> None:
        """窗口显示时初始化按钮位置 + 自动刷新设备列表."""
        super().showEvent(event)
        QTimer.singleShot(0, self._reposition_edge_buttons)
        if not hasattr(self, "_startup_refreshed"):
            self._startup_refreshed = True
            QTimer.singleShot(300, self._startup_load_devices)

    def _startup_load_devices(self) -> None:
        """启动时自动加载设备列表并连接MCGS设备"""
        self._refresh_device_list()
        QTimer.singleShot(1000, self._auto_connect_mcgs_devices)

    def _auto_connect_mcgs_devices(self) -> None:
        """自动连接所有MCGS设备"""
        if self._mcgs_controller is None:
            return
        try:
            device_ids = self._mcgs_controller.list_devices()
            if device_ids:
                logger.info("[自动连接] 准备连接 %d 个MCGS设备: %s", len(device_ids), device_ids)
                for device_id in device_ids:
                    self._mcgs_controller.connect_device(device_id)
                logger.info("[自动连接] 已发送连接请求")
        except Exception as e:
            logger.error("[自动连接] 失败: %s", e)

    def _create_left_panel(self) -> QWidget:
        left_widget = QWidget()
        # 增加左侧面板最小宽度，确保5列设备列表和4个按钮完整显示
        left_widget.setMinimumWidth(380)
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(12, 10, 10, 8)
        left_layout.setSpacing(6)
        left_widget.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)
        # 浅色背景
        left_widget.setStyleSheet("background-color: #FFFFFF;")

        # ── 标题行 ──
        title_layout = QHBoxLayout()
        title_layout.setSpacing(8)

        # 添加折叠按钮到标题左侧
        self._collapse_btn = QPushButton("<")
        self._collapse_btn.setObjectName("left_collapse_btn")
        self._collapse_btn.setFixedSize(24, 24)
        self._collapse_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._collapse_btn.setToolTip("折叠面板")
        # 使用安全的 ASCII 字符确保跨平台兼容性
        try:
            from ui.design_tokens import DT

            collapse_font = DT.T.get_font(*DT.T.BODY_SMALL)
            self._collapse_btn.setFont(collapse_font)
        except (ImportError, AttributeError):
            from PySide6.QtGui import QFont

            self._collapse_btn.setFont(QFont("Segoe UI Symbol", 10))
        title_layout.addWidget(self._collapse_btn)
        self._collapse_btn.clicked.connect(self._collapse_left_panel)

        self._left_title_label = QLabel(TextConstants.DEVICE_LIST_TITLE)
        try:
            self._left_title_label.setFont(QFont("Inter", 20, QFont.Weight.Bold))
        except NameError:
            from PySide6.QtGui import QFont

            self._left_title_label.setFont(QFont("Segoe UI Variable", 20, QFont.Weight.Bold))
        self._left_title_label.setStyleSheet("color: #24292F; background: transparent;")
        title_layout.addWidget(self._left_title_label)
        title_layout.addStretch()

        left_layout.addLayout(title_layout)

        # ── 搜索框 ──
        self._search_edit = LineEdit(TextConstants.SEARCH_PLACEHOLDER)
        self._search_edit.setMinimumHeight(30)
        self._search_edit.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self._search_edit.textChanged.connect(self._filter_devices)
        left_layout.addWidget(self._search_edit)

        # ── 分组/扁平切换 ──
        group_toggle_layout = QHBoxLayout()
        group_toggle_layout.setContentsMargins(4, 4, 4, 0)
        self._group_toggle = QPushButton("☰ 按类型分组")
        self._group_toggle.setFixedHeight(24)
        self._group_toggle.setCheckable(True)
        self._group_toggle.setCursor(Qt.CursorShape.PointingHandCursor)
        self._group_toggle.setStyleSheet(
            """
            QPushButton {
                background: transparent; border: 1px solid #D0D7DE; border-radius: 4px;
                padding: 2px 8px; font-size: 11px; color: #57606A; text-align: left;
            }
            QPushButton:checked {
                background: #E3F2FD; border-color: #0969DA; color: #0969DA; font-weight: 600;
            }
            QPushButton:hover { background: #F3F4F6; }
        """
        )
        self._group_toggle.clicked.connect(self._toggle_group_by_type)
        group_toggle_layout.addWidget(self._group_toggle)
        group_toggle_layout.addStretch()
        left_layout.addLayout(group_toggle_layout)

        # ── 设备树 (弹性拉伸) ──
        self._device_tree = DeviceTree(self)
        self._device_tree.currentItemChanged.connect(self._on_device_selected)
        self._device_tree.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self._device_tree._context_menu_handler = self._handle_device_context_menu
        left_layout.addWidget(self._device_tree, 1)

        # ── 底部按钮行 (等距排满栏目宽度，不同背景色) ──
        from ui.widgets import PrimaryButton, SecondaryButton, Colors

        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(4)
        btn_layout.setContentsMargins(0, 12, 0, 12)

        # 添加设备按钮 - 绿色背景
        self._add_device_btn = QPushButton(TextConstants.ADD_DEVICE_BTN)
        self._add_device_btn.setFixedHeight(36)
        self._add_device_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._add_device_btn.setStyleSheet(
            """
            QPushButton { background-color: #2DA44E; color: white; border: 1px solid #1a7a37; border-radius: 6px; padding: 5px 12px; font-size: 12px; font-weight: 500; }
            QPushButton:hover { background-color: #54AE76; }
            QPushButton:pressed { background-color: #1a7a37; }
        """
        )
        self._add_device_btn.clicked.connect(self._add_device)
        self._add_device_btn.setToolTip("添加新设备")
        btn_layout.addWidget(self._add_device_btn, 1)

        # 设备配置按钮 - 蓝色背景
        self._device_config_btn = QPushButton("设备配置")
        self._device_config_btn.setFixedHeight(36)
        self._device_config_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._device_config_btn.setStyleSheet(
            """
            QPushButton { background-color: #0969DA; color: white; border: 1px solid #0550AE; border-radius: 6px; padding: 5px 12px; font-size: 12px; font-weight: 500; }
            QPushButton:hover { background-color: #4CA1ED; }
            QPushButton:pressed { background-color: #0550AE; }
        """
        )
        self._device_config_btn.clicked.connect(self._on_show_mcgsm_config)
        self._device_config_btn.setToolTip("打开MCGS设备可视化配置编辑器")
        btn_layout.addWidget(self._device_config_btn, 1)

        # 刷新列表按钮 - 灰色背景
        self._refresh_btn = QPushButton("刷新列表")
        self._refresh_btn.setFixedHeight(36)
        self._refresh_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._refresh_btn.setStyleSheet(
            """
            QPushButton { background-color: #6B7280; color: white; border: 1px solid #4B5563; border-radius: 6px; padding: 5px 12px; font-size: 12px; font-weight: 500; }
            QPushButton:hover { background-color: #9CA3AF; }
            QPushButton:pressed { background-color: #4B5563; }
        """
        )
        self._refresh_btn.clicked.connect(self._on_full_refresh_devices)
        self._refresh_btn.setToolTip("刷新设备列表，显示最新的设备配置")
        btn_layout.addWidget(self._refresh_btn, 1)

        # 扫描网络按钮 - 橙色背景
        self._scan_btn = QPushButton("扫描设备")
        self._scan_btn.setFixedHeight(36)
        self._scan_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._scan_btn.setStyleSheet(
            """
            QPushButton { background-color: #F59E0B; color: white; border: 1px solid #D97706; border-radius: 6px; padding: 5px 12px; font-size: 12px; font-weight: 500; }
            QPushButton:hover { background-color: #FBBF24; }
            QPushButton:pressed { background-color: #D97706; }
        """
        )
        self._scan_btn.clicked.connect(self._open_scan_dialog)
        self._scan_btn.setToolTip("扫描局域网内的 Modbus TCP 设备")
        btn_layout.addWidget(self._scan_btn, 1)

        # 删除设备按钮 - 红色背景
        self._remove_btn = QPushButton(TextConstants.REMOVE_DEVICE_BTN)
        self._remove_btn.setFixedHeight(36)
        self._remove_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._remove_btn.setStyleSheet(
            """
            QPushButton { background-color: #CF222E; color: white; border: 1px solid #A51D26; border-radius: 6px; padding: 5px 12px; font-size: 12px; font-weight: 500; }
            QPushButton:hover { background-color: #E85B65; }
            QPushButton:pressed { background-color: #A51D26; }
        """
        )
        self._remove_btn.clicked.connect(self._remove_device)
        self._remove_btn.setToolTip("删除选中的设备（需确认）")
        btn_layout.addWidget(self._remove_btn, 1)

        left_layout.addLayout(btn_layout)

        return left_widget

    def _create_welcome_page(self) -> QWidget:
        welcome_page = QWidget()
        welcome_layout = QVBoxLayout(welcome_page)
        welcome_layout.setContentsMargins(40, 40, 40, 40)
        welcome_layout.setSpacing(20)

        welcome_label = QLabel(TextConstants.WELCOME_TITLE)
        welcome_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        welcome_label.setFont(QFont("Inter", 24, QFont.Weight.Bold))
        welcome_label.setStyleSheet("color: #24292F;")
        welcome_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        welcome_layout.addWidget(welcome_label)

        welcome_sub_label = QLabel(TextConstants.WELCOME_SUBTITLE)
        welcome_sub_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        welcome_sub_label.setFont(QFont("Inter", 14))
        welcome_sub_label.setStyleSheet("color: #57606A;")
        welcome_sub_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        welcome_layout.addWidget(welcome_sub_label)
        welcome_layout.addStretch()

        return welcome_page

    def _create_monitor_page(self) -> QWidget:
        page = self._monitor_controller.build(
            parent=self,
            styles=AppStyles.__dict__,
            constants=TextConstants.__dict__,
            on_expand_panel=self._expand_left_panel,
        )

        self._expand_btn = self._monitor_controller.expand_btn
        self._device_title_label = self._monitor_controller.device_title_label
        self._device_name_label = self._monitor_controller.device_name_label
        self._device_desc_label = self._monitor_controller.device_desc_label
        self._last_update_label = self._monitor_controller.last_update_label
        self._device_status_badge = self._monitor_controller.device_status_badge
        self._right_splitter = self._monitor_controller.right_splitter
        self._monitor_tabs = self._monitor_controller.monitor_tabs
        self._data_cards_layout = self._monitor_controller.data_cards_layout
        self._chart_layout = self._monitor_controller.chart_layout
        self._register_table = self._monitor_controller.get_register_table()

        return page

    def _update_cards_display(self) -> None:
        from ui.widgets import DataCard

        self._monitor_controller.update_cards_display(
            current_device_id=self._current_device_id or "",
            device_cards=self._device_cards,
            data_card_cls=DataCard,
        )

    def _log_message(self, message: str, level: str = "INFO", device_id: str = None, operation: str = None) -> None:
        if hasattr(self, "_monitor_controller") and self._monitor_controller:
            self._monitor_controller.append_log(message, level)

    def _connect_signals(self) -> None:
        """连接设备管理器信号到UI处理槽函数（v4.0 DataBus订阅版）"""
        self._device_manager.device_added.connect(self._on_device_added)
        self._device_manager.device_removed.connect(self._on_device_removed)
        self._device_manager.device_connected.connect(self._on_device_connected)
        self._device_manager.device_disconnected.connect(self._on_device_disconnected)
        self._device_manager.device_error.connect(self._on_device_error)

        if hasattr(self._device_manager, "async_poll_success"):
            self._device_manager.async_poll_success.connect(self._on_async_poll_success)
            self._device_manager.async_poll_failed.connect(self._on_async_poll_failed)
            self._device_manager.async_poll_timeout.connect(self._on_async_poll_timeout)

        self._write_manager.confirm_required.connect(self._show_write_confirmation)
        self._write_manager.operation_result.connect(self._on_write_operation_result)

        self._connect_data_bus()

        # 将 Python logging 模块的输出也转发到 UI 日志
        self._install_log_handler()

    def _install_log_handler(self) -> None:
        """安装 Python logging 处理器，将所有日志记录转发到 UI 日志框"""
        import logging

        class _UiLogHandler(logging.Handler):
            def __init__(self, append_fn):
                super().__init__()
                self._append_fn = append_fn

            def emit(self, record):
                level = (
                    "ERROR"
                    if record.levelno >= logging.ERROR
                    else (
                        "WARNING"
                        if record.levelno >= logging.WARNING
                        else "INFO" if record.levelno >= logging.INFO else "DEBUG"
                    )
                )
                try:
                    self._append_fn(record.getMessage(), level)
                except Exception:
                    pass

        handler = _UiLogHandler(self._log_message)
        handler.setLevel(logging.INFO)
        logging.getLogger().addHandler(handler)

    def _connect_data_bus(self):
        """订阅 DataBus 全局信号 — 统一数据更新入口（规范控制点⑥）"""
        from core.foundation.data_bus import DataBus

        bus = DataBus.instance()

        bus.subscribe("device_data_updated", self._on_bus_device_data_updated)
        bus.subscribe("comm_error", self._on_bus_comm_error)
        bus.subscribe("device_status_changed", self._on_bus_device_status_changed)
        bus.subscribe("device_connected", self._on_bus_device_connected)
        bus.subscribe("device_disconnected", self._on_bus_device_disconnected)
        bus.subscribe("alarm_triggered", self._on_bus_alarm_triggered)

    @Slot(str, dict)
    def _on_bus_device_data_updated(self, device_id: str, data: dict):
        """DataBus 数据更新回调 — 更新日志 + 监控面板数据"""
        if not data:
            return

        is_current_device = device_id == self._current_device_id
        is_mcgs = self._is_mcgs_device(device_id)

        if is_current_device:
            self._last_update_label.setText(f"{TextConstants.LAST_UPDATE_LABEL} {datetime.now().strftime('%H:%M:%S')}")

        was_none = self._current_device_id is None
        should_update = False

        if was_none:
            self._current_device_id = device_id
            should_update = True
        elif is_current_device:
            should_update = True

        if should_update:
            if not self._device_cards.get(device_id):
                card_configs = self._build_cards_config_from_json(device_id, data)
                if card_configs:
                    self._device_cards[device_id] = card_configs
                    self._update_cards_display()
            elif not self._monitor_controller.get_all_card_widgets():
                self._update_cards_display()

            if is_mcgs:
                return
            self._update_card_values(data)
            self._update_register_table_from_data(data)

    def _is_mcgs_device(self, device_id: str) -> bool:
        if self._mcgs_controller is None:
            return False
        reader = self._mcgs_controller.get_reader()
        if reader is None:
            return False
        try:
            return device_id in (reader.list_devices() or [])
        except Exception:
            return False

    def _auto_create_cards_from_data(self, device_id: str, data: dict):
        """根据接收到的数据自动创建数据卡片配置"""
        cards_config = []
        point_decimal_map = {}
        point_unit_map = {}
        point_desc_map = {}

        # 直接从 devices.json 读取点位描述/单位/精度
        try:
            import json as _json
            from pathlib import Path as _Path

            cfg_path = _Path(__file__).parent.parent / "config" / "devices.json"
            if cfg_path.exists():
                with open(cfg_path, "r", encoding="utf-8") as f:
                    cfg_data = _json.load(f)
                for dev in cfg_data.get("devices", []):
                    if dev.get("id") == device_id:
                        for pt in dev.get("points", []):
                            pname = pt.get("name", "")
                            if pt.get("description"):
                                point_desc_map[pname] = pt["description"]
                            if pt.get("unit"):
                                point_unit_map[pname] = pt["unit"]
                            dp = pt.get("decimal_places")
                            if dp is not None:
                                point_decimal_map[pname] = int(dp)
                        break
        except Exception:
            pass

        # 备用：从MCGS读取器配置获取（如果JSON读取失败）
        try:
            if self._mcgs_controller and (not point_desc_map or not point_unit_map):
                reader = self._mcgs_controller.get_reader()
                if reader:
                    dev_cfg = reader.get_device_config(device_id)
                    if dev_cfg and dev_cfg.points:
                        for p in dev_cfg.points:
                            if p.name not in point_desc_map and p.description:
                                point_desc_map[p.name] = p.description
                            if p.name not in point_unit_map and p.unit:
                                point_unit_map[p.name] = p.unit
                            if p.name not in point_decimal_map:
                                point_decimal_map[p.name] = p.decimal_places
        except Exception:
            pass

        if not data:
            keys = [pn for pn in point_desc_map] or [pn for pn in point_unit_map]
        else:
            keys = list(data.keys())

        for key in keys:
            display_name = point_desc_map.get(key, key)
            unit = point_unit_map.get(key, "")

            cards_config.append(
                {
                    "title": display_name,
                    "register_name": key,
                    "description": display_name if display_name != key else "",
                    "unit": unit,
                    "decimal_places": point_decimal_map.get(key, 2),
                }
            )

        if cards_config:
            self._device_cards[device_id] = cards_config
            self._update_cards_display()
            logger.info(f"[Monitor] 自动为设备 {device_id} 创建了 {len(cards_config)} 个数据卡片")

    _json_cards_cache: Dict[str, tuple] = {}

    def _build_cards_config_from_json(self, device_id: str, data: dict) -> list:
        """直接从 devices.json 读取点位描述/单位/精度，构建卡片配置（带内存缓存）"""
        if device_id in MainWindow._json_cards_cache:
            desc_map, unit_map, dec_map = MainWindow._json_cards_cache[device_id]
        else:
            desc_map, unit_map, dec_map = {}, {}, {}
            try:
                import json as _json
                from pathlib import Path as _Path

                cfg_path = _Path(__file__).parent.parent / "config" / "devices.json"
                if cfg_path.exists():
                    with open(cfg_path, "r", encoding="utf-8") as f:
                        cfg = _json.load(f)
                    for dev in cfg.get("devices", []):
                        if dev.get("id") == device_id:
                            for pt in dev.get("points", []):
                                pname = pt.get("name", "")
                                if pt.get("description"):
                                    desc_map[pname] = pt["description"]
                                if pt.get("unit"):
                                    unit_map[pname] = pt["unit"]
                                dp = pt.get("decimal_places")
                                if dp is not None:
                                    dec_map[pname] = int(dp)
                            break
                    MainWindow._json_cards_cache[device_id] = (desc_map, unit_map, dec_map)
            except Exception:
                pass

        cards = []
        for key in data.keys():
            display_name = desc_map.get(key, key)
            cards.append(
                {
                    "title": display_name,
                    "register_name": key,
                    "description": display_name,
                    "unit": unit_map.get(key, ""),
                    "decimal_places": dec_map.get(key, 2),
                }
            )
        return cards

    def _refresh_cards_for_device(self, device_id: str) -> None:
        """强制刷新指定设备的数据卡片（标记待刷新，下次数据到达时重建）"""
        if not device_id:
            return
        # 标记该设备需要在下一次数据更新时重建卡片
        if not hasattr(self, "_pending_card_refresh"):
            self._pending_card_refresh = set()
        self._pending_card_refresh.add(device_id)
        logger.info(f"[Monitor] 标记设备 {device_id} 卡片待刷新")

    def _update_register_table_from_data(self, data: dict):
        """根据DataBus数据更新寄存器表格"""
        if self._register_table is None:
            return
        try:
            _ = self._register_table.rowCount()
        except RuntimeError:
            return
        self._register_table.setRowCount(len(data))
        COL_COUNT = 3  # Function Code, Variable Name, Value
        self._register_table.setColumnCount(COL_COUNT)
        for row, (name, value_info) in enumerate(data.items()):
            if isinstance(value_info, dict):
                value = value_info.get("value", "")
            else:
                value = str(value_info) if value_info is not None else ""
            self._register_table.setItem(row, 0, QTableWidgetItem("03"))
            self._register_table.setItem(row, 1, QTableWidgetItem(name))
            self._register_table.setItem(row, 2, QTableWidgetItem(str(value)))
            for col in range(COL_COUNT):
                item = self._register_table.item(row, col)
                if item:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self._register_table.resizeColumnsToContents()

    @Slot(str, str)
    def _on_bus_comm_error(self, device_id: str, error_msg: str):
        """DataBus 通信错误回调"""
        logger.warning("[DataBus] 通信错误 [%s]: %s", device_id, error_msg)
        self._status_msg_label.setText(f"通信错误: {device_id} - {error_msg}")

    @Slot(str, str)
    def _on_bus_device_status_changed(self, device_id: str, status: str):
        """DataBus 设备状态变更回调"""
        logger.info("[DataBus] 设备状态 [%s]: %s", device_id, status)
        self._update_status_bar()

    @Slot(str)
    def _on_bus_device_connected(self, device_id: str):
        """DataBus 设备连接回调 — 更新状态并强制刷新设备树按钮"""
        self._update_device_status_in_tree(device_id, is_connected=True)
        self._log_message("设备已连接", "SUCCESS", device_id, "CONNECT")
        # 延迟刷新设备树确保按钮状态正确（使用50ms避免与_toggle_mcgs_connection的500ms竞态）
        QTimer.singleShot(50, lambda: self._refresh_device_list(self._search_edit.text()))

    @Slot(str)
    def _on_bus_device_disconnected(self, device_id: str):
        """DataBus 设备断开回调 — 更新状态并强制刷新设备树按钮"""
        self._update_device_status_in_tree(device_id, is_connected=False)
        QTimer.singleShot(50, lambda: self._refresh_device_list(self._search_edit.text()))
        self._log_message("设备已断开", "INFO", device_id, "DISCONNECT")

    @Slot(str, str, str, float)
    def _on_bus_alarm_triggered(self, device_id: str, param_name: str, alarm_type: str, value: float):
        """DataBus 报警触发回调"""
        self._log_message(f"报警: {param_name} {alarm_type} (值={value})", "WARNING", device_id, "ALARM")

    def _apply_theme(self) -> None:
        self._theme_manager.apply_theme()

    def _on_full_refresh_devices(self) -> None:
        """刷新设备列表（重载 MCGS 配置 + 刷新 UI）"""
        # 重新创建 MCGS 读取器以加载最新的 devices.json
        if self._mcgs_controller:
            self._mcgs_controller.reset_reader()
            self._get_or_create_mcgsm_reader()

        self._refresh_device_list()
        self._log_message("设备列表已刷新", "INFO")

    def _refresh_device_list(self, search_text: str = "") -> None:
        """从 MCGS 配置 (devices.json) 加载设备列表，替代通用协议栈"""
        if self._mcgs_controller is not None and self._mcgs_controller.get_reader() is None:
            try:
                self._get_or_create_mcgsm_reader()
            except Exception:
                pass

        if self._mcgs_controller is not None and self._mcgs_controller.get_reader() is not None:
            try:
                from pathlib import Path

                config_path = Path(__file__).parent.parent / "config" / "devices.json"
                if config_path.exists():
                    self._mcgs_controller.get_reader().load_config(config_path)
            except Exception as e:
                logger.warning("重新加载设备配置失败: %s", e)

        current_device_id = self._current_device_id

        self._device_tree.currentItemChanged.disconnect(self._on_device_selected)
        self._device_tree.clear()

        # ── 从 MCGS Controller 获取设备列表 ──
        mcgs_devices = self._get_mcgs_device_list()

        if self._group_by_type:
            self._build_device_tree_grouped(mcgs_devices, search_text)
        else:
            self._build_device_tree_flat(mcgs_devices, search_text)

        # 重新连接信号
        self._device_tree.currentItemChanged.connect(self._on_device_selected)

        # 恢复之前选中的设备
        if current_device_id:
            restored = self._restore_selected_device(current_device_id)
            if not restored and self._device_tree.topLevelItemCount() > 0:
                self._restore_first_available_child()
        elif self._device_tree.topLevelItemCount() > 0:
            self._restore_first_available_child()

        # 刷新后自适应调整尺寸
        QTimer.singleShot(0, self._update_tree_adaptive_sizes)

    def _update_device_status_in_tree(self, device_id: str, is_connected: bool) -> None:
        """更新设备树中指定设备的状态显示和操作按钮（不重建整树）"""
        status_text = "在线" if is_connected else "离线"
        for i in range(self._device_tree.topLevelItemCount()):
            item = self._device_tree.topLevelItem(i)
            if item is None:
                continue
            # 扁平模式：直接检查
            if item.data(0, Qt.ItemDataRole.UserRole) == device_id:
                item.setText(3, status_text)
                self._attach_action_widget(item, device_id, is_connected)
                return
            # 分组模式：检查子节点
            for ci in range(item.childCount()):
                child = item.child(ci)
                if child and child.data(0, Qt.ItemDataRole.UserRole) == device_id:
                    child.setText(3, status_text)
                    self._attach_action_widget(child, device_id, is_connected)
                    return

    def _build_device_tree_flat(self, mcgs_devices: list, search_text: str = "") -> None:
        """构建扁平设备树（每台设备一行）"""
        for dev in mcgs_devices:
            device_id = dev["id"]
            name = dev["name"]

            if search_text and search_text.lower() not in name.lower():
                continue

            is_connected = dev.get("connected", False)
            device_type = dev.get("device_type", "") or "MCGS触摸屏"

            item = QTreeWidgetItem()
            item.setText(0, device_type)
            item.setText(1, str(device_id))
            item.setText(2, str(dev.get("point_count", 0)))
            item.setText(3, "在线" if is_connected else "离线")

            item.setTextAlignment(0, Qt.AlignmentFlag.AlignCenter)
            item.setTextAlignment(1, Qt.AlignmentFlag.AlignCenter)
            item.setTextAlignment(2, Qt.AlignmentFlag.AlignCenter)
            item.setTextAlignment(3, Qt.AlignmentFlag.AlignCenter)
            item.setData(0, Qt.ItemDataRole.UserRole, device_id)
            item.setSizeHint(0, QSize(0, 48))
            self._device_tree.addTopLevelItem(item)
            self._attach_action_widget(item, device_id, is_connected)

    def _build_device_tree_grouped(self, mcgs_devices: list, search_text: str = "") -> None:
        """构建分组设备树（按 device_type 分组）"""
        from collections import OrderedDict

        groups: Dict[str, list] = OrderedDict()
        for dev in mcgs_devices:
            dtype = dev.get("device_type", "") or "未分类"
            if dtype not in groups:
                groups[dtype] = []
            groups[dtype].append(dev)

        for dtype, dev_list in groups.items():
            # 过滤：至少有一个设备的名称匹配搜索词
            filtered = [d for d in dev_list if not search_text or search_text.lower() in d["name"].lower()]
            if not filtered:
                continue

            # 创建分组父节点
            group_item = QTreeWidgetItem()
            group_item.setText(0, f"{dtype} ({len(filtered)} 设备)")
            group_item.setText(1, "")
            group_item.setText(2, "")
            group_item.setText(3, "")
            group_item.setFirstColumnSpanned(True)
            group_item.setSizeHint(0, QSize(0, 36))
            group_item.setFlags(group_item.flags() & ~Qt.ItemFlag.ItemIsSelectable)
            # 展开状态
            group_item.setExpanded(True)
            self._device_tree.addTopLevelItem(group_item)

            for dev in filtered:
                device_id = dev["id"]
                name = dev["name"]
                is_connected = dev.get("connected", False)

                child = QTreeWidgetItem()
                child.setText(0, name)
                child.setText(1, str(device_id))
                child.setText(2, str(dev.get("point_count", 0)))
                child.setText(3, "在线" if is_connected else "离线")

                child.setTextAlignment(0, Qt.AlignmentFlag.AlignCenter)
                child.setTextAlignment(1, Qt.AlignmentFlag.AlignCenter)
                child.setTextAlignment(2, Qt.AlignmentFlag.AlignCenter)
                child.setTextAlignment(3, Qt.AlignmentFlag.AlignCenter)
                child.setData(0, Qt.ItemDataRole.UserRole, device_id)
                child.setSizeHint(0, QSize(0, 48))

                group_item.addChild(child)
                self._attach_action_widget(child, device_id, is_connected)

    def _attach_action_widget(self, item: QTreeWidgetItem, device_id: str, is_connected: bool) -> None:
        """为树节点附加连接/断开按钮（自动替换旧按钮）"""
        # 先移除旧按钮
        old_widget = self._device_tree.itemWidget(item, 4)
        if old_widget:
            self._device_tree.removeItemWidget(item, 4)
            old_widget.deleteLater()

        action_widget = QWidget()
        action_widget.setStyleSheet("background: transparent;")
        action_layout = QHBoxLayout(action_widget)
        action_layout.setContentsMargins(4, 4, 4, 4)
        action_layout.setSpacing(6)

        if is_connected:
            conn_btn = DangerButton(TextConstants.BTN_DISCONNECT)
            conn_btn.setToolTip("断开连接")
        else:
            conn_btn = SuccessButton(TextConstants.BTN_CONNECT)
            conn_btn.setToolTip("连接设备")
        conn_btn.setMinimumHeight(30)
        conn_btn.setMaximumHeight(38)
        conn_btn.setMinimumWidth(44)
        conn_btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        conn_btn.clicked.connect(lambda checked, did=device_id: self._toggle_mcgs_connection(did))

        action_layout.addWidget(conn_btn)
        self._device_tree.setItemWidget(item, 4, action_widget)

    def _restore_selected_device(self, device_id: str) -> bool:
        """恢复选中的设备（支持扁平/分组两种模式）"""
        for i in range(self._device_tree.topLevelItemCount()):
            item = self._device_tree.topLevelItem(i)
            if item is None:
                continue
            # 扁平模式：直接顶级
            if item.data(0, Qt.ItemDataRole.UserRole) == device_id:
                self._device_tree.setCurrentItem(item)
                return True
            # 分组模式：遍历子节点
            for ci in range(item.childCount()):
                child = item.child(ci)
                if child and child.data(0, Qt.ItemDataRole.UserRole) == device_id:
                    self._device_tree.setCurrentItem(child)
                    return True
        return False

    def _restore_first_available_child(self) -> None:
        """选中第一个可用设备（分组模式下跳过父节点）"""
        for i in range(self._device_tree.topLevelItemCount()):
            item = self._device_tree.topLevelItem(i)
            if item is None:
                continue
            # 如果是父节点且有子节点，选第一个子节点
            if item.childCount() > 0:
                self._device_tree.setCurrentItem(item.child(0))
                return
            # 扁平模式：直接选
            if item.data(0, Qt.ItemDataRole.UserRole) is not None:
                self._device_tree.setCurrentItem(item)
                return

    def _get_mcgs_device_list(self) -> list:
        """
        从 MCGS 配置获取设备列表数据。

        优先从 MCGSModbusReader 获取，备用从 DeviceManager 网关模型获取。

        Returns:
            list[dict]: 每个元素包含 id, name, ip, port, point_count, connected 等字段
        """
        result = []

        if self._mcgs_controller is not None:
            reader = self._mcgs_controller.get_reader()
            if reader is not None:
                try:
                    device_ids = reader.list_devices()
                    for device_id in device_ids:
                        config = reader.get_device_config(device_id)
                        if config:
                            # ★ 关键修复：使用 device_id（与 DataBus 发布一致）
                            # 而不是 config.id（可能不同）
                            result.append(
                                {
                                    "id": device_id,  # 使用原始 device_id，确保与 DataBus 一致
                                    "name": getattr(config, "name", device_id) or device_id,
                                    "device_type": getattr(config, "device_type", ""),
                                    "ip": getattr(config, "ip", ""),
                                    "port": getattr(config, "port", 0),
                                    "point_count": len(getattr(config, "points", [])),
                                    "connected": reader.is_device_connected(device_id),
                                }
                            )
                except Exception as e:
                    logger.warning("获取 MCGS 设备列表失败: %s", e)

        if not result and hasattr(self._device_manager, "get_all_gateway_models"):
            try:
                models = self._device_manager.get_all_gateway_models()
                for gw_id, model in models.items():
                    result.append(
                        {
                            "id": gw_id,  # 使用网关ID
                            "name": model.name,
                            "ip": model.ip,
                            "port": model.port,
                            "point_count": model.variable_count,
                            "connected": self._device_manager.is_gateway_connected(gw_id),
                        }
                    )
            except Exception as e:
                logger.warning("从网关模型获取设备列表失败: %s", e)

        return result

    def _toggle_mcgs_connection(self, device_id: str) -> None:
        """切换 MCGS 设备的连接状态（通过 MCGSController 异步调度）"""
        if self._mcgs_controller is None:
            self._log_message("MCGS 控制器未初始化", "ERROR")
            return

        reader = self._mcgs_controller.get_reader()
        is_connected = reader and reader.is_device_connected(device_id)

        if is_connected:
            self._mcgs_controller.disconnect_device(device_id)
            self._log_message(f"MCGS 断开请求已发送: {device_id}", "INFO", device_id, "DISCONNECT")
        else:
            self._mcgs_controller.connect_device(device_id)
            self._log_message(f"MCGS 连接请求已发送: {device_id}", "INFO", device_id, "CONNECT")

    def _filter_devices(self) -> None:
        search_text = self._search_edit.text()
        self._refresh_device_list(search_text)

    def _toggle_group_by_type(self) -> None:
        """切换设备树分组/扁平模式"""
        self._group_by_type = self._group_toggle.isChecked()
        self._refresh_device_list(self._search_edit.text())

    def _update_tree_adaptive_sizes(self) -> None:
        """根据面板当前宽度自适应调整设备树行高和操作列按钮.

        当面板被 splitter 拉伸/压缩时调用,
        动态调整行高和操作列宽度以保持比例协调。
        """
        if self._left_panel_collapsed:
            return

        panel_width = self._left_panel.width()
        # 行高随面板宽度线性缩放: 460px→48, 650px→56, 850px→64
        row_height = int(40 + max(0, min(panel_width - 460, 500)) * 0.032)
        row_height = max(48, min(row_height, 70))

        # 操作列宽度: 面板越宽, 操作按钮可分配越多空间
        # 基准 150px (两按钮各50px + 间距6px + 边距)，最大 200px
        action_col_width = int(130 + max(0, panel_width - 600) * 0.15)
        action_col_width = max(150, min(action_col_width, 200))  # 从170-260调整为150-200
        # 操作列索引为4，因为前面有4列（0-3）
        self._device_tree.setColumnWidth(4, action_col_width)

        # 更新所有行的行高和操作列按钮
        from PySide6.QtWidgets import QPushButton

        btn_min_h = max(26, row_height - 20)
        btn_max_h = row_height - 10

        for i in range(self._device_tree.topLevelItemCount()):
            item = self._device_tree.topLevelItem(i)
            item.setSizeHint(0, QSize(0, row_height))
            # 更新操作列中的按钮尺寸，操作列索引为4
            action_widget = self._device_tree.itemWidget(item, 4)
            if action_widget:
                for btn in action_widget.findChildren(QPushButton):
                    btn.setMinimumHeight(btn_min_h)
                    btn.setMaximumHeight(btn_max_h)

    def _add_device(self) -> None:
        from ui.device_type_dialogs import AddMCGSDeviceDialog
        import json

        logger.debug("Opening add MCGS device dialog")
        try:
            dialog = AddMCGSDeviceDialog(self)
            if dialog.exec():
                config = dialog.get_device_config()
                if config:
                    from pathlib import Path

                    devices_path = Path("config/devices.json")
                    data = {}
                    if devices_path.exists():
                        with open(devices_path, "r", encoding="utf-8") as f:
                            data = json.load(f)

                    for existing in data.get("devices", []):
                        if existing.get("id") == config["id"]:
                            QMessageBox.warning(self, "错误", f"设备ID '{config['id']}' 已存在")
                            return

                    data.setdefault("devices", []).append(config)
                    data.setdefault("_meta", {})["version"] = "3.0.4"
                    data["_meta"]["source"] = "AddMCGSDeviceDialog"

                    with open(devices_path, "w", encoding="utf-8") as f:
                        json.dump(data, f, ensure_ascii=False, indent=2)

                    self._status_msg_label.setText(f"MCGS设备添加成功: {config['id']}")
                    MainWindow._json_cards_cache.clear()
                    self._refresh_device_list("")
                    self._update_status_bar()
                    logger.info("[MCGS] 设备已添加: %s @ %s:%s", config["id"], config["ip"], config["port"])
        except Exception as e:
            logger.error("添加MCGS设备失败: %s", str(e))
            QMessageBox.warning(self, "添加设备失败", f"无法添加设备: {str(e)}")
            self._status_msg_label.setText(f"添加设备失败: {str(e)}")

    def _remove_device(self) -> None:
        item = self._device_tree.currentItem()
        if not item:
            QMessageBox.warning(self, UIMessages.SELECT_DEVICE_TITLE, UIMessages.SELECT_DEVICE_MSG)
            return

        device_id = item.data(0, Qt.ItemDataRole.UserRole)
        self._remove_device_by_id(device_id)

    def _remove_device_by_id(self, device_id: str) -> None:
        reply = QMessageBox.question(
            self,
            UIMessages.CONFIRM_DELETE_TITLE,
            UIMessages.CONFIRM_DELETE_MSG,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                is_mcgs = False

                if self._mcgs_controller:
                    reader = self._mcgs_controller.get_reader()
                    if reader and device_id in reader.list_devices():
                        is_mcgs = True

                if is_mcgs:
                    success = self._remove_mcgs_device(device_id)
                else:
                    success = self._device_manager.remove_device(device_id)

                if success:
                    logger.info(LogMessages.DEVICE_REMOVED.format(device_id=device_id))
                    self._status_msg_label.setText(f"设备移除成功: {device_id}")

                    if device_id in self._device_cards:
                        del self._device_cards[device_id]

                    was_current = self._current_device_id == device_id
                    if was_current:
                        self._current_device_id = None
                        self._stack_widget.setCurrentIndex(0)

                    self._refresh_device_list(self._search_edit.text())
                else:
                    logger.error("移除设备失败: %s", device_id)
                    QMessageBox.warning(self, "移除设备失败", f"无法移除设备: {device_id}")
                    self._status_msg_label.setText(f"移除设备失败: {device_id}")
            except Exception as e:
                logger.error("移除设备失败: %s", str(e))
                QMessageBox.warning(self, "移除设备失败", f"无法移除设备: {str(e)}")
                self._status_msg_label.setText(f"移除设备失败: {str(e)}")

    def _remove_mcgs_device(self, device_id: str) -> bool:
        """删除MCGS设备：断开连接 + 从配置文件移除"""
        try:
            if self._mcgs_controller:
                if device_id in self._mcgs_controller._polling_device_ids:
                    self._mcgs_controller.disconnect_device(device_id)

                reader = self._mcgs_controller.get_reader()
                if reader and device_id in reader._devices:
                    del reader._devices[device_id]

                import json
                from pathlib import Path

                config_path = Path(__file__).parent.parent / "config" / "devices.json"
                if config_path.exists():
                    with open(config_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    devices = data.get("devices", [])
                    original_count = len(devices)
                    data["devices"] = [d for d in devices if d.get("id") != device_id]
                    if len(data["devices"]) < original_count:
                        with open(config_path, "w", encoding="utf-8") as f:
                            json.dump(data, f, ensure_ascii=False, indent=2)
                        logger.info("[MCGS删除] 已从配置文件移除设备: %s", device_id)

                return True
        except Exception as e:
            logger.error("[MCGS删除] 删除设备异常: %s - %s", device_id, e)
            return False

    def _set_operation_lock(self, locked: bool, btn: Optional[QPushButton] = None, loading_text: str = "") -> None:
        """设置操作锁，防止重复点击（P0-2修复）

        Args:
            locked: 是否锁定
            btn: 需要禁用的按钮（可选）
            loading_text: 按钮在锁定时显示的文字
        """
        self._operation_lock = locked

        if locked and btn is not None:
            # 锁定：禁用按钮并显示加载状态
            self._current_operation_btn = btn
            original_text = btn.text()
            btn.setEnabled(False)
            btn.setText(loading_text or original_text)
            # 存储原始文字以便恢复
            if not hasattr(btn, "_original_text"):
                btn._original_text = original_text
        elif not locked and self._current_operation_btn is not None:
            # 解锁：恢复按钮状态
            btn = self._current_operation_btn
            btn.setEnabled(True)
            if hasattr(btn, "_original_text"):
                btn.setText(btn._original_text)
                delattr(btn, "_original_text")
            self._current_operation_btn = None

    def _connect_device_by_id(self, device_id: str) -> None:
        # P0-2修复：检查操作锁，防止重复点击
        if self._operation_lock:
            logger.debug("连接操作正在进行中，忽略重复请求")
            return

        try:
            # 获取当前行的连接按钮并锁定
            current_item = self._device_tree.currentItem()
            connect_btn = None
            if current_item:
                action_widget = self._device_tree.itemWidget(current_item, 4)
                if action_widget:
                    for btn in action_widget.findChildren(QPushButton):
                        if "连接" in btn.text() or "Connect" in btn.text():
                            connect_btn = btn
                            break

            self._set_operation_lock(True, connect_btn, "连接中...")
            self._status_msg_label.setText("正在连接...")

            success, error_type, error_msg = self._device_manager.connect_device(device_id)
            if success:
                logger.info(LogMessages.DEVICE_CONNECTING.format(device_id=device_id))
                self._status_msg_label.setText(f"设备连接成功: {device_id}")
            else:
                # 根据错误类型生成更详细的错误消息
                detailed_msg = f"{error_msg}\n\n错误类型: {error_type}\n\n是否需要重新配置设备?"
                reply = QMessageBox.warning(
                    self,
                    UIMessages.CONNECT_FAILED_TITLE,
                    detailed_msg,
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                )

                # 如果用户选择"是"，则打开设备编辑对话框
                if reply == QMessageBox.StandardButton.Yes:
                    self._edit_device_by_id(device_id)

                logger.error(LogMessages.DEVICE_CONNECT_FAILED.format(device_id=device_id))
                logger.error("连接失败详情: 类型=%s, 消息=%s", error_type, error_msg)
                self._status_msg_label.setText(f"设备连接失败: {error_msg}")
        except Exception as e:
            logger.error("连接设备失败: %s", str(e))
            QMessageBox.warning(self, UIMessages.CONNECT_FAILED_TITLE, f"连接设备时发生错误: {str(e)}")
            self._status_msg_label.setText(f"连接设备失败: {str(e)}")
        finally:
            # P0-2修复：解锁操作
            self._set_operation_lock(False)

            # 刷新设备列表，更新连接状态
            self._refresh_device_list()
            # 更新状态栏显示
            self._update_status_bar()

    def _disconnect_device_by_id(self, device_id: str) -> None:
        # P0-2修复：检查操作锁，防止重复点击
        if self._operation_lock:
            logger.debug("断开操作正在进行中，忽略重复请求")
            return

        try:
            # 获取当前行的断开按钮并锁定
            current_item = self._device_tree.currentItem()
            disconnect_btn = None
            if current_item:
                action_widget = self._device_tree.itemWidget(current_item, 4)
                if action_widget:
                    for btn in action_widget.findChildren(QPushButton):
                        if "断开" in btn.text() or "Disconnect" in btn.text():
                            disconnect_btn = btn
                            break

            self._set_operation_lock(True, disconnect_btn, "断开中...")
            success = self._device_manager.disconnect_device(device_id)
            if success:
                logger.info(LogMessages.DEVICE_DISCONNECTED.format(device_id=device_id))
                self._status_msg_label.setText(f"设备断开成功: {device_id}")
            else:
                logger.error("断开设备失败: %s", device_id)
                self._status_msg_label.setText(f"断开设备失败: {device_id}")
        except Exception as e:
            logger.error("断开设备失败: %s", str(e))
            self._status_msg_label.setText(f"断开设备失败: {str(e)}")
        finally:
            # P0-2修复：解锁操作
            self._set_operation_lock(False)

            # 刷新设备列表，更新连接状态
            self._refresh_device_list()
            # 更新状态栏显示
            self._update_status_bar()

    def _toggle_connection_by_id(self, device_id: str) -> None:
        device = self._device_manager.get_device(device_id)
        if device:
            # 自动选中当前设备（如果未选中）
            if self._current_device_id != device_id:
                for i in range(self._device_tree.topLevelItemCount()):
                    item = self._device_tree.topLevelItem(i)
                    if item and item.data(0, Qt.ItemDataRole.UserRole) == device_id:
                        self._device_tree.setCurrentItem(item)
                        break

            if device.get_status() == DeviceStatus.CONNECTED:
                self._disconnect_device_by_id(device_id)
            else:
                self._connect_device_by_id(device_id)

    def _edit_device(self) -> None:
        item = self._device_tree.currentItem()
        if not item:
            QMessageBox.warning(self, UIMessages.SELECT_DEVICE_TITLE, UIMessages.SELECT_DEVICE_MSG)
            return

        device_id = item.data(0, Qt.ItemDataRole.UserRole)
        self._edit_device_by_id(device_id)

    def _edit_device_by_id(self, device_id: str) -> None:
        device = self._device_manager.get_device(device_id)
        if not device:
            return

        try:
            from ui.device_type_dialogs import DeviceTypeManager

            config = device.get_device_config()
            device_type_manager = DeviceTypeManager("device_types.json")
            dialog = AddDeviceDialog(device_type_manager, self, edit_mode=True, device_config=config)
            if dialog.exec():
                new_config = dialog.get_device_config()
                success = self._device_manager.edit_device(device_id, new_config)
                if success:
                    logger.info(LogMessages.DEVICE_EDIT_SUCCESS.format(device_id=device_id))
                    self._status_msg_label.setText(f"设备编辑成功: {device_id}")
                else:
                    logger.error("编辑设备失败: %s", device_id)
                    self._status_msg_label.setText(f"编辑设备失败: {device_id}")
        except Exception as e:
            logger.error("编辑设备失败: %s", str(e))
            QMessageBox.warning(self, "编辑设备失败", f"无法编辑设备: {str(e)}")
            self._status_msg_label.setText(f"编辑设备失败: {str(e)}")
        finally:
            # 确保对话框被正确释放
            if "dialog" in locals():
                dialog.deleteLater()
            if "device_type_manager" in locals():
                del device_type_manager

    def _handle_device_context_menu(self, action_type: str, target_id: str) -> None:
        """处理设备树右键菜单操作"""
        if action_type == "edit":
            self._edit_mcgs_device(target_id)
        elif action_type == "copy":
            self._copy_mcgs_device(target_id)
        elif action_type == "delete":
            self._remove_device_by_id(target_id)
        elif action_type == "scan":
            self._open_scan_dialog()

    def _edit_mcgs_device(self, device_id: str) -> None:
        """编辑 MCGS 设备的连接参数（IP/端口/从站ID 等）"""
        logger.info("[编辑设备] 开始编辑设备: %s", device_id)
        import json
        from pathlib import Path
        from ui.device_type_dialogs import EditMCGSDeviceDialog

        devices_path = Path("config/devices.json")
        if not devices_path.exists():
            QMessageBox.warning(self, "错误", "设备配置文件不存在")
            return

        try:
            with open(devices_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            device_cfg = None
            for dev in data.get("devices", []):
                if dev.get("id") == device_id:
                    device_cfg = dev
                    break

            if not device_cfg:
                QMessageBox.warning(self, "错误", f"未找到设备: {device_id}")
                return

            logger.info("[编辑设备] 找到设备配置: %s", device_cfg.get("name", ""))

            was_connected = False
            if self._mcgs_controller:
                reader = self._mcgs_controller.get_reader()
                if reader and hasattr(reader, "is_device_connected") and reader.is_device_connected(device_id):
                    was_connected = True
                    logger.info("[编辑设备] 设备正在连接，先断开: %s", device_id)
                    self._mcgs_controller.disconnect_device(device_id)

            logger.info("[编辑设备] 打开编辑对话框...")
            dialog = EditMCGSDeviceDialog(device_cfg, parent=self)
            result = dialog.exec()
            logger.info("[编辑设备] 对话框关闭，结果: %s", result)

            if result == QDialog.DialogCode.Accepted:
                new_config = dialog.get_device_config()
                if new_config:
                    for i, dev in enumerate(data.get("devices", [])):
                        if dev.get("id") == device_id:
                            data["devices"][i] = new_config
                            break

                    with open(devices_path, "w", encoding="utf-8") as f:
                        json.dump(data, f, ensure_ascii=False, indent=2)

                    logger.info(
                        "[MCGS] 设备连接参数已更新: %s -> %s:%s",
                        device_id,
                        new_config["ip"],
                        new_config["port"],
                    )
                    self._status_msg_label.setText(f"[OK] 设备参数已更新: {new_config['ip']}:{new_config['port']}")

                    if self._mcgs_controller:
                        self._mcgs_controller.reset_reader()
                        if was_connected:
                            QTimer.singleShot(300, lambda did=device_id: self._mcgs_controller.connect_device(did))

                    self._refresh_device_list(self._search_edit.text())
        except Exception as e:
            logger.error("编辑 MCGS 设备失败: %s", str(e))
            QMessageBox.warning(self, "编辑失败", f"无法编辑设备: {str(e)}")

    def _copy_mcgs_device(self, device_id: str) -> None:
        """复制已有 MCGS 设备，生成新 ID 和默认名称"""
        import json
        from pathlib import Path
        from ui.device_type_dialogs import AddMCGSDeviceDialog

        devices_path = Path("config/devices.json")
        if not devices_path.exists():
            return

        try:
            with open(devices_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            source_cfg = None
            for dev in data.get("devices", []):
                if dev.get("id") == device_id:
                    source_cfg = dev
                    break

            if not source_cfg:
                return

            dialog = AddMCGSDeviceDialog(self)
            dialog.type_combo.setCurrentIndex(0)
            copy_idx = dialog.copy_combo.findData(source_cfg)
            if copy_idx > 0:
                dialog.copy_combo.setCurrentIndex(copy_idx)
                dialog._on_copy_changed(copy_idx)

            base_name = source_cfg.get("name", device_id)
            dialog.device_name_edit.setText(f"{base_name}_副本")
            base_id = device_id
            dialog.device_id_edit.setText(f"{base_id}_copy")

            if dialog.exec():
                config = dialog.get_device_config()
                if config:
                    data.setdefault("devices", []).append(config)
                    data["_meta"]["source"] = "CopyMCGSDevice"
                    with open(devices_path, "w", encoding="utf-8") as f:
                        json.dump(data, f, ensure_ascii=False, indent=2)
                    self._refresh_device_list("")
                    self._log_message(f"设备已复制: {device_id} -> {config['id']}", "INFO")
        except Exception as e:
            logger.error("复制设备失败: %s", str(e))
            QMessageBox.warning(self, "复制失败", f"无法复制设备: {str(e)}")

    @Slot()
    def _open_scan_dialog(self) -> None:
        """打开 Modbus TCP 局域网设备扫描对话框"""
        dialog = DeviceScanDialog(parent=self)
        dialog.device_added.connect(lambda cfg: self._on_scan_device_added(cfg))
        result = dialog.exec()

        if result == QDialog.DialogCode.Accepted or dialog.get_added_devices():
            added_count = len(dialog.get_added_devices())
            if added_count > 0:
                self._log_message(f"扫描添加了 {added_count} 个新设备", "INFO")
                self._on_full_refresh_devices()

    def _on_scan_device_added(self, device_cfg: dict) -> None:
        """扫描发现设备被添加后的回调"""
        logger.info(
            "[Scan] 新设备已通过扫描添加: %s @ %s:%s",
            device_cfg.get("id"),
            device_cfg.get("ip"),
            device_cfg.get("port"),
        )

    def _update_register_table(self, registers: List) -> None:
        self._register_table.setRowCount(len(registers))

        for row, reg in enumerate(registers):
            self._register_table.setItem(row, 0, QTableWidgetItem(str(reg.get("address", ""))))
            self._register_table.setItem(row, 1, QTableWidgetItem(str(reg.get("function_code", ""))))
            self._register_table.setItem(row, 2, QTableWidgetItem(reg.get("name", "")))
            self._register_table.setItem(row, 3, QTableWidgetItem(str(reg.get("value", ""))))
            self._register_table.setItem(row, 4, QTableWidgetItem(reg.get("unit", "")))

            for col in range(5):
                item = self._register_table.item(row, col)
                if item:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

        self._register_table.resizeColumnsToContents()

    def _show_device_type_dialog(self) -> None:
        from ui.device_type_dialogs import DeviceTypeDialog

        logger.debug("Opening MCGS device type management dialog")
        dialog = DeviceTypeDialog(self)
        dialog.exec()
        self._refresh_device_list()
        self._update_status_bar()

    def _show_batch_operations(self) -> None:
        dialog = BatchOperationsDialog(self._device_manager, self)
        dialog.operations_completed.connect(self._on_batch_operations_completed)
        dialog.exec()

    def _on_batch_operations_completed(self, success_count: int, total_count: int) -> None:
        self._refresh_device_list()
        self._update_status_bar()
        logger.info(LogMessages.BATCH_OPS_COMPLETE.format(success=success_count, total=total_count))

    def _show_export_device_config_dialog(self) -> None:
        """打开设备配置导出对话框"""
        file_path, _ = QFileDialog.getSaveFileName(self, "导出设备配置", "", "JSON Files (*.json)")

        if file_path:
            # 导出所有设备配置
            success = self._device_manager.export_devices_config(file_path)
            if success:
                QMessageBox.information(
                    self,
                    UIMessages.EXPORT_SUCCESS_TITLE,
                    UIMessages.DEVICE_CONFIG_EXPORT_SUCCESS_MSG.format(path=file_path),
                )
                logger.info(LogMessages.DATA_EXPORT_SUCCESS.format(path=file_path))
            else:
                QMessageBox.warning(self, UIMessages.EXPORT_FAILED_TITLE, UIMessages.DEVICE_CONFIG_EXPORT_FAILED_MSG)

    def _show_import_device_config_dialog(self) -> None:
        """打开设备配置导入对话框"""
        file_path, _ = QFileDialog.getOpenFileName(self, "导入设备配置", "", "JSON Files (*.json)")

        if file_path:
            # 确认导入
            reply = QMessageBox.question(
                self,
                UIMessages.IMPORT_CONFIRM_TITLE,
                UIMessages.DEVICE_CONFIG_IMPORT_CONFIRM_MSG,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )

            if reply == QMessageBox.StandardButton.Yes:
                # 导入设备配置
                success = self._device_manager.import_devices_config(file_path, overwrite=True)
                if success:
                    QMessageBox.information(
                        self, UIMessages.IMPORT_SUCCESS_TITLE, UIMessages.DEVICE_CONFIG_IMPORT_SUCCESS_MSG
                    )
                    # 刷新设备列表
                    self._refresh_device_list()
                else:
                    QMessageBox.warning(
                        self, UIMessages.IMPORT_FAILED_TITLE, UIMessages.DEVICE_CONFIG_IMPORT_FAILED_MSG
                    )

    def _set_app_icon(self):
        from PySide6.QtGui import QPixmap, QPainter, QColor, QFont
        from PySide6.QtCore import Qt

        pixmap = QPixmap(64, 64)
        pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.setBrush(QColor("#3B82F6"))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawRoundedRect(4, 4, 56, 56, 12, 12)
        painter.setPen(QColor("white"))
        font = QFont("Segoe UI", 28, QFont.Weight.Bold)
        painter.setFont(font)
        painter.drawText(pixmap.rect(), Qt.AlignmentFlag.AlignCenter, "M")
        painter.end()
        self.setWindowIcon(QPixmap(pixmap))

    def _show_about(self) -> None:
        dialog = QDialog(self)
        dialog.setWindowTitle(UIMessages.ABOUT_TITLE)
        dialog.setMinimumWidth(480)
        layout = QVBoxLayout(dialog)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)

        info_label = QLabel(UIMessages.ABOUT_MSG)
        info_label.setWordWrap(True)
        info_label.setTextFormat(Qt.TextFormat.RichText)
        info_label.setOpenExternalLinks(True)
        info_label.setStyleSheet("font-size: 13px;")
        layout.addWidget(info_label)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        license_btn = QPushButton("许可证 (&L)")
        license_btn.setStyleSheet(
            "QPushButton { padding: 6px 16px; border: 1px solid #D0D0D0; border-radius: 4px; }"
            "QPushButton:hover { background: #F0F0F0; }"
        )
        license_btn.clicked.connect(self._show_license)
        btn_layout.addWidget(license_btn)

        ok_btn = QPushButton("确定")
        ok_btn.setDefault(True)
        ok_btn.setStyleSheet(
            "QPushButton { padding: 6px 20px; background: #3B82F6; color: white; "
            "border: none; border-radius: 4px; }"
            "QPushButton:hover { background: #2563EB; }"
        )
        ok_btn.clicked.connect(dialog.accept)
        btn_layout.addWidget(ok_btn)

        layout.addLayout(btn_layout)
        dialog.exec()

    def _show_license(self) -> None:
        try:
            license_path = Path(__file__).parent.parent / "LICENSE"
            if license_path.exists():
                text = license_path.read_text(encoding="utf-8")
            else:
                text = "许可证文件未找到。"
        except Exception as e:
            text = f"无法读取许可证文件: {e}"

        dialog = QDialog(self)
        dialog.setWindowTitle("许可证 — MIT License")
        dialog.setMinimumSize(560, 440)
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(16, 16, 16, 16)

        text_edit = QPlainTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setPlainText(text)
        text_edit.setStyleSheet("font-family: Consolas, monospace; font-size: 13px;")
        layout.addWidget(text_edit)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        ok_btn = QPushButton("确定")
        ok_btn.setDefault(True)
        ok_btn.clicked.connect(dialog.accept)
        btn_layout.addWidget(ok_btn)
        layout.addLayout(btn_layout)

        dialog.exec()

    def _on_device_selected(self, current: Optional[QTreeWidgetItem], previous: Optional[QTreeWidgetItem]) -> None:
        if not current:
            self._stack_widget.setCurrentIndex(0)
            return

        device_id = current.data(0, Qt.ItemDataRole.UserRole)

        if device_id == self._current_device_id:
            if self._stack_widget.currentIndex() != 1:
                self._stack_widget.setCurrentIndex(1)
            if not self._monitor_controller.get_all_card_widgets():
                if self._device_cards.get(device_id):
                    self._monitor_controller.clear_all_cards()
                    self._update_cards_display()
            return

        logger.info("[设备选中] _on_device_selected 触发: device_id=%s (之前=%s)", device_id, self._current_device_id)
        self._current_device_id = device_id

        is_mcgs_device = False
        config = None
        reader = None
        if self._mcgs_controller:
            reader = self._mcgs_controller.get_reader()
        if reader:
            config = reader.get_device_config(device_id)
            if config:
                is_mcgs_device = True
                is_connected = (
                    reader.is_device_connected(device_id) if hasattr(reader, "is_device_connected") else False
                )
                self._log_message(
                    f"MCGS设备已选中: {config.name} ({config.ip}:{config.port})", "INFO", device_id, "SELECT"
                )
                self._monitor_controller.clear_all_cards()
                self._update_monitor_page_for_mcgs(device_id, config, is_connected=is_connected)
            else:
                if device_id in (reader.list_devices() or []):
                    logger.warning("[设备选中] 设备在reader列表中但get_config返回None，尝试重建: %s", device_id)
                    is_mcgs_device = True
                    self._monitor_controller.clear_all_cards()
                    self._auto_create_cards_from_data(device_id, {})
                    self._device_name_label.setText(device_id)
                    self._device_desc_label.setText("")
                    self._device_desc_label.hide()
                    self._device_status_badge.status = "offline"
                    self._last_update_label.setText("-")
                    self._stack_widget.setCurrentIndex(1)
                else:
                    self._log_message(f"设备已选中: {device_id}", "INFO", device_id, "SELECT")
        else:
            self._log_message(f"设备已选中: {device_id}", "INFO", device_id, "SELECT")

        if not is_mcgs_device:
            self._stack_widget.setCurrentIndex(0)

    def _update_monitor_page_for_mcgs(self, device_id: str, config, is_connected: bool = True) -> None:
        """为MCGS设备更新监控面板"""
        self._stack_widget.setCurrentIndex(1)
        display_name = getattr(config, "name", device_id) or device_id
        self._device_name_label.setText(display_name)
        description = getattr(config, "description", "") or ""
        if description:
            self._device_desc_label.setText(f"({description})")
            self._device_desc_label.show()
        else:
            self._device_desc_label.setText("")
            self._device_desc_label.hide()

        if is_connected:
            self._device_status_badge.status = "online"
        else:
            self._device_status_badge.status = "offline"
        self._last_update_label.setText(datetime.now().strftime("%H:%M:%S") if is_connected else "-")
        points = getattr(config, "points", []) or []

        logger.info("[MCGS监控页] 创建/重建卡片: %s (已连接=%s)", device_id, is_connected)
        register_map = []
        cards_config = []
        for p in points:
            p_name = getattr(p, "name", "")
            p_addr = getattr(p, "addr", "")
            p_unit = getattr(p, "unit", "")
            p_desc = getattr(p, "description", "")
            register_map.append(
                {
                    "address": p_addr,
                    "function_code": "03",
                    "name": p_name,
                    "value": "-",
                    "unit": p_unit,
                }
            )
            cards_config.append(
                {
                    "title": p_desc if p_desc else p_name,
                    "register_name": p_name,
                    "description": p_desc,
                    "unit": p_unit,
                    "decimal_places": getattr(p, "decimal_places", 2),
                }
            )
        self._device_cards[device_id] = cards_config
        self._update_cards_display()
        self._update_register_table(register_map)

    @Slot(str)
    def _on_device_added(self, device_id: str) -> None:
        self._refresh_device_list(self._search_edit.text())
        self._update_status_bar()
        message = f"设备已添加"
        self._status_msg_label.setText(f"设备已添加: {device_id}")
        self._log_message(message, "SUCCESS", device_id, "ADD")

    @Slot(str)
    def _on_device_removed(self, device_id: str) -> None:
        self._refresh_device_list(self._search_edit.text())
        self._update_status_bar()
        self._stack_widget.setCurrentIndex(0)
        message = f"设备已移除"
        self._status_msg_label.setText(f"设备已移除: {device_id}")
        self._log_message(message, "INFO", device_id, "REMOVE")

    @Slot(str)
    def _on_device_connected(self, device_id: str) -> None:
        self._update_device_status_in_tree(device_id, is_connected=True)
        self._update_status_bar()
        message = f"设备已连接"
        self._status_msg_label.setText(f"设备已连接: {device_id}")
        logger.info(LogMessages.DEVICE_CONNECTED.format(device_id=device_id))
        self._log_message(message, "SUCCESS", device_id, "CONNECT")

    @Slot(str)
    def _on_device_disconnected(self, device_id: str) -> None:
        self._update_device_status_in_tree(device_id, is_connected=False)
        self._update_status_bar()
        message = f"设备已断开"
        self._status_msg_label.setText(f"设备已断开: {device_id}")
        logger.info(LogMessages.DEVICE_DISCONNECTED.format(device_id=device_id))
        self._log_message(message, "INFO", device_id, "DISCONNECT")
        if self._current_device_id == device_id:
            try:
                reader = None
                if self._mcgs_controller:
                    reader = self._mcgs_controller.get_reader()
                config = reader.get_device_config(device_id) if reader else None
                if config is not None:
                    self._update_monitor_page_for_mcgs(device_id, config, is_connected=False)
                else:
                    self._current_device_id = None
                    self._stack_widget.setCurrentIndex(0)
            except Exception as e:
                logger.warning("断开时更新监控页异常: %s", e)
                self._current_device_id = None
                self._stack_widget.setCurrentIndex(0)

    @Slot(str, dict)
    def _on_device_data_updated(self, device_id: str, data: dict) -> None:
        """设备数据更新回调 — 已迁移至 DataBus 订阅（保留向后兼容）"""
        self._on_bus_device_data_updated(device_id, data)

    def _update_card_values(self, data: Dict[str, Any]) -> None:
        """更新数据卡片数值 + 收集曲线历史数据"""
        try:
            updated_count = 0
            card_widgets = self._monitor_controller.get_all_card_widgets()

            for register_name, card in card_widgets.items():
                if not isinstance(card, DataCard):
                    continue
                if register_name and register_name in data:
                    value_info = data[register_name]

                    if isinstance(value_info, dict):
                        value = value_info.get("value", 0)
                        unit = value_info.get("unit", "")
                    elif isinstance(value_info, (int, float)):
                        value = float(value_info)
                        unit = ""
                    elif isinstance(value_info, str):
                        # 尝试从格式化字符串（如"36.00 ℃"）中提取数值和单位
                        match = re.match(r"([-+]?\d*\.?\d+)\s*(.*)", value_info.strip())
                        if match:
                            value = float(match.group(1))
                            unit = match.group(2).strip()  # 提取单位（如"℃"）
                        else:
                            # 非数值字符串（如"gg"），直接显示原始值
                            card.set_value(value_info.strip())
                            updated_count += 1
                            continue
                    else:
                        value = 0.0
                        unit = ""

                    precision = self._monitor_controller.get_card_precision(register_name)
                    display_str = f"{round(value, precision):.{precision}f}"
                    full_display = f"{display_str} {unit}" if unit else display_str
                    card.set_value(full_display)
                    updated_count += 1
                    if hasattr(self._monitor_controller, "update_selected_card_chart"):
                        self._monitor_controller.update_selected_card_chart(register_name, float(value))

            if updated_count > 0:
                logger.debug("[Card-UPDATE] 成功更新 %d/%d 个卡片", updated_count, len(card_widgets))
        except Exception as e:
            logger.error(f"[Card-UPDATE] 异常: {e}", exc_info=True)

    @Slot(str, str)
    def _on_device_error(self, device_id: str, error: str) -> None:
        logger.error("Device error: %s - %s", device_id, error)
        message = f"错误: {error}"
        self._status_msg_label.setText(f"设备错误: {device_id} - {error}")
        self._update_status_bar()
        self._log_message(message, "ERROR", device_id, "ERROR")

    # ══════════════════════════════════════════════
    # 异步轮询结果处理（v3.0新增）
    # ══════════════════════════════════════════════

    @Slot(str, dict, float)
    def _on_async_poll_success(
        self,
        device_id: str,
        data: dict,
        response_time_ms: float,
    ) -> None:
        """
        处理异步轮询成功（带性能数据）

        此方法在主线程中被调用（Qt Signal/Slot自动排队），
        可以安全地更新UI元素。

        Args:
            device_id: 设备唯一标识符
            data: 轮询到的数据字典
            response_time_ms: 响应耗时（毫秒），用于性能监控
        """
        # 性能监控：如果响应时间过长，记录警告
        if response_time_ms > 100:
            logger.warning(
                "[PERF] 设备 %s 轮询响应时间过长: %.1fms (正常<50ms)",
                device_id,
                response_time_ms,
            )

        # 更新状态栏显示（带性能数据）
        self._status_msg_label.setText(f"设备 {device_id[:8]}... 轮询完成 ({response_time_ms:.0f}ms)")
        self._update_status_bar()

    @Slot(str, str, str)
    def _on_async_poll_failed(
        self,
        device_id: str,
        error_type: str,
        error_msg: str,
    ) -> None:
        """处理异步轮询失败"""
        logger.warning(
            "Async poll failed: %s - [%s] %s",
            device_id,
            error_type,
            error_msg,
        )

        # 更新状态栏（不阻塞UI，因为这是异步通知）
        self._status_msg_label.setText(f"轮询异常: {device_id[:8]}... - {error_type}")
        self._update_status_bar()

    @Slot(str, float)
    def _on_async_poll_timeout(
        self,
        device_id: str,
        elapsed_ms: float,
    ) -> None:
        """处理异步轮询超时"""
        logger.debug("Async poll timeout: %s (%.1fms)", device_id, elapsed_ms)

        # 超时是常见情况（Modbus RTU慢速设备），仅记录不告警
        if elapsed_ms > 200:
            logger.info(
                "设备 %s 轮询超时较严重: %.1fms",
                device_id,
                elapsed_ms,
            )

    def _update_status_bar(self) -> None:
        """更新状态栏信息"""
        mcgs_online = 0
        mcgs_offline = 0
        mcgs_error = 0
        mcgs_total = 0

        if self._mcgs_controller:
            try:
                reader = self._mcgs_controller.get_reader()
                if reader:
                    device_ids = reader.list_devices()
                    mcgs_total = len(device_ids)
                    for did in device_ids:
                        if reader.is_device_connected(did):
                            mcgs_online += 1
                        else:
                            mcgs_offline += 1
            except Exception as e:
                logger.warning("更新状态栏异常: %s", e)

        total_count = mcgs_total
        online_count = mcgs_online
        offline_count = mcgs_offline
        error_count = mcgs_error

        self._status_total_label.setText(f"设备 {total_count}")
        self._status_online_label.setText(f"● 在线 {online_count}")
        self._status_offline_label.setText(f"● 离线 {offline_count}")
        self._status_error_label.setText(f"● 错误 {error_count}")

        current_time = datetime.now().strftime("%H:%M:%S")
        self._status_time_label.setText(f"更新时间: {current_time}")

    def cleanup(self) -> None:
        if getattr(self, "_cleaned", False):
            return
        self._cleaned = True
        logger.info(LogMessages.APP_SHUTDOWN)
        self._save_ui_preferences()

        try:
            if self._mcgs_controller is not None:
                try:
                    self._mcgs_controller.device_data_updated.disconnect(self._on_mcgsm_raw_data_updated)
                except Exception:
                    pass
                try:
                    self._mcgs_controller.device_error.disconnect(self._on_mcgsm_device_error)
                except Exception:
                    pass
                try:
                    self._mcgs_controller.polling_started.disconnect(self._on_mcgsm_polling_started)
                except Exception:
                    pass
                try:
                    self._mcgs_controller.polling_stopped.disconnect(self._on_mcgsm_polling_stopped)
                except Exception:
                    pass
                try:
                    self._mcgs_controller.poll_cycle_completed.disconnect(self._on_mcgsm_poll_cycle_completed)
                except Exception:
                    pass
                self._mcgs_controller.cleanup()
                logger.info("MCGSController 已清理")
        except Exception as e:
            logger.warning("MCGSController 清理异常: %s", e)

        try:
            from core.foundation.data_bus import DataBus

            bus = DataBus.instance()
            bus.unsubscribe("device_data_updated", self._on_bus_device_data_updated)
            bus.unsubscribe("comm_error", self._on_bus_comm_error)
            bus.unsubscribe("device_status_changed", self._on_bus_device_status_changed)
            bus.unsubscribe("device_connected", self._on_bus_device_connected)
            bus.unsubscribe("device_disconnected", self._on_bus_device_disconnected)
            bus.unsubscribe("alarm_triggered", self._on_bus_alarm_triggered)
            logger.info("DataBus 订阅已释放")
        except Exception as e:
            logger.warning("DataBus 取消订阅异常: %s", e)

        try:
            from ui.animation_scheduler import AnimationScheduler

            AnimationScheduler.cleanup_instance()
        except Exception:
            pass

        if hasattr(self, "_cleanup_scheduler") and self._cleanup_scheduler:
            self._cleanup_scheduler.stop()
        self._device_manager.cleanup()

    def _load_ui_preferences(self) -> None:
        try:
            self._device_cards = {}
            self._device_charts = {}
            panel_state = self._ui_prefs.load_panel_state()
            if panel_state:
                self._left_panel_collapsed = panel_state.get("left_collapsed", False)
                self._left_panel_saved_size = panel_state.get("left_width", 480)

            raw_cards = getattr(self._ui_prefs, "_data", {}).get("device_cards", {})
            for device_id in raw_cards:
                self._device_cards[device_id] = self._ui_prefs.load_cards(device_id)
            raw_charts = getattr(self._ui_prefs, "_data", {}).get("device_charts", {})
            for device_id in raw_charts:
                self._device_charts[device_id] = self._ui_prefs.load_charts(device_id)

            logger.debug(
                "UI 偏好配置已加载: %d 设备卡片, %d 设备图表",
                len(self._device_cards),
                len(self._device_charts),
            )
        except Exception:
            logger.exception("加载 UI 偏好配置失败")

    def _save_ui_preferences(self) -> None:
        try:
            for device_id, cards in self._device_cards.items():
                self._ui_prefs.save_cards(device_id, cards)
            for device_id, charts in self._device_charts.items():
                self._ui_prefs.save_charts(device_id, charts)
            self._ui_prefs.save_panel_state(
                left_collapsed=self._left_panel_collapsed,
                left_width=self._left_panel_saved_size,
                right_width=self.width() - self._left_panel_saved_size,
            )
            logger.debug("UI 偏好配置已保存")
        except Exception:
            logger.exception("保存 UI 偏好配置失败")

    # ══════════════════════════════════════════════════════════
    # ✅ 新增：权限管理相关方法
    # ══════════════════════════════════════════════════════════

    def _on_permission_changed(self) -> None:
        """
        权限状态变化槽函数

        当用户登录/登出或角色变更时调用，
        更新UI状态（按钮启用/禁用、菜单显示等）。
        """
        self._update_permission_ui()

    def _on_session_timeout(self, username: str) -> None:
        """
        会话超时槽函数

        Args:
            username: 超时的用户名
        """
        from PySide6.QtWidgets import QMessageBox

        QMessageBox.warning(
            self,
            "会话超时",
            f"用户 '{username}' 的会话已超时（30分钟无操作），已自动登出。\n\n" "请重新登录以继续操作。",
        )

        # 更新状态栏
        self._status_msg_label.setText(f"会话超时: {username} 已自动登出")

        self._update_permission_ui()

    def _update_permission_ui(self) -> None:
        """
        根据当前权限状态更新UI

        - 更新状态栏显示
        - 启用/禁用写操作相关按钮
        - 更新菜单项可见性
        """
        if not hasattr(self, "_permission_mgr"):
            return

        if self._permission_mgr.is_logged_in:
            # 已登录状态
            username = self._permission_mgr.current_username
            role = self._permission_mgr.current_user_role

            if role:
                role_display = role.display_name
                can_write = self._permission_mgr.can_write

                # 更新状态栏（如果有专门的权限标签）
                if hasattr(self, "_permission_label"):
                    self._permission_label.setText(f"用户: {username} ({role_display})")

                logger.debug("权限UI更新 [用户=%s, 角色=%s, 可写=%s]", username, role_display, can_write)
        else:
            # 未登录状态
            if hasattr(self, "_permission_label"):
                self._permission_label.setText("未登录")

            logger.debug("权限UI更新: 未登录")

    def show_login_dialog(self) -> None:
        """显示登录对话框"""
        from ui.login_dialog import LoginDialog

        dialog = LoginDialog(self._permission_mgr, parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            username = dialog.username
            role = self._permission_mgr.current_user_role

            self._status_msg_label.setText(f"登录成功: {username} ({role.display_name if role else ''})")
            logger.info("用户登录成功: %s", username)
        else:
            logger.debug("登录对话框被取消")

    def show_logout_action(self) -> None:
        """执行登出操作"""
        if not self._permission_mgr.is_logged_in:
            return

        from PySide6.QtWidgets import QMessageBox

        reply = QMessageBox.question(
            self,
            "确认登出",
            "确定要退出当前账户吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )

        if reply == QMessageBox.StandardButton.Yes:
            username = self._permission_mgr.current_username
            self._permission_mgr.logout()

            self._status_msg_label.setText(f"已登出: {username}")
            logger.info("用户已登出: %s", username)

    # ═══════════════════════════════════════════════════════════
    # ✅ MCGS 快速连接模块 (v2.0 - 2026-04-23)
    # ═══════════════════════════════════════════════════════════

    def _init_mcgsm_controller(self) -> None:
        """初始化MCGS控制器（异步调度+Signal中转，替代直接调用通信层）"""
        try:
            from ui.controllers.mcgs_controller import MCGSController
            from core.services.mcgs_service import MCGSService
            from core.services.history_service import HistoryService

            # 使用绝对路径避免运行时CWD不一致导致数据库文件错位
            from pathlib import Path as _Path

            _project_root = _Path(__file__).resolve().parent.parent
            _raw_db_path = (
                getattr(self._db_manager, "_db_path", None) or self._db_manager.db_path
                if hasattr(self._db_manager, "db_path")
                else "data/equipment_management.db"
            )
            db_path = str(_Path(_raw_db_path))
            if not _Path(db_path).is_absolute():
                db_path = str(_project_root / _raw_db_path)

            logger.info("HistoryService 数据库路径: %s", db_path)
            history_service = HistoryService(db_path=db_path)
            history_service.initialize()

            mcgs_service = MCGSService(
                history_service=history_service,
            )

            self._mcgs_controller = MCGSController(mcgs_service=mcgs_service, parent=self)

            # 数据更新通过 DataBus 订阅接收（见 _connect_data_bus）
            self._mcgs_controller.device_error.connect(self._on_mcgsm_device_error)
            self._mcgs_controller.polling_started.connect(self._on_mcgsm_polling_started)
            self._mcgs_controller.polling_stopped.connect(self._on_mcgsm_polling_stopped)
            self._mcgs_controller.poll_cycle_completed.connect(self._on_mcgsm_poll_cycle_completed)
            self._mcgs_controller.device_data_updated.connect(self._on_mcgsm_raw_data_updated)

            logger.info("MCGSController 初始化完成")

        except Exception as e:
            logger.error(f"MCGSController初始化失败: {e}")
            self._mcgs_controller = None
            QMessageBox.warning(self, "MCGS模块初始化失败", f"MCGS控制器初始化失败:\n{str(e)}")

    def _get_or_create_mcgsm_reader(self):
        """获取或创建MCGS读取器（通过Controller延迟初始化+缓存）"""
        if self._mcgs_controller is None:
            return None

        reader = self._mcgs_controller.get_reader()
        if reader is not None:
            return reader

        try:
            from pathlib import Path

            config_path = Path(__file__).parent.parent / "config" / "devices.json"

            if config_path.exists():
                reader = self._mcgs_controller.create_reader(str(config_path))
                if reader:
                    logger.info(f"MCGS读取器已创建: {config_path}")
                    return reader
                return None
            else:
                QMessageBox.warning(
                    self,
                    "配置文件缺失",
                    f"MCGS配置文件不存在:\n{config_path}\n\n" f"请创建 devices.json 并配置设备参数。",
                )
                return None

        except Exception as e:
            logger.error(f"创建MCGS读取器失败: {e}")
            QMessageBox.critical(self, "MCGS连接错误", f"无法初始化MCGS读取器:\n{str(e)}")
            return None

    @Slot(str, str)
    def _on_mcgsm_device_error(self, device_id: str, error_msg: str):
        """MCGS设备错误回调"""
        logger.warning("[%s] MCGS错误: %s", device_id, error_msg)
        self._status_msg_label.setText(f"MCGS错误 [{device_id}]: {error_msg}")
        self._log_message(f"MCGS错误 [{device_id}]: {error_msg}", "ERROR", device_id, "ERROR")
        # 更新受影响的设备状态（只更新文本，不重建整树）
        self._update_device_status_in_tree(device_id, is_connected=False)

    @Slot(str, dict)
    def _on_mcgsm_raw_data_updated(self, device_id: str, data: dict):
        if device_id == self._current_device_id and data:
            self._update_card_values(data)

    @Slot()
    def _on_mcgsm_polling_started(self):
        """轮询启动回调"""
        logger.info("MCGS自动轮询已启动")
        self._log_message("MCGS自动轮询已启动", "INFO")

    @Slot()
    def _on_mcgsm_polling_stopped(self):
        """轮询停止回调"""
        logger.info("MCGS自动轮询已停止")
        self._log_message("MCGS自动轮询已停止", "INFO")

    @Slot(int, int)
    def _on_mcgsm_poll_cycle_completed(self, success_count: int, fail_count: int):
        """轮询周期完成回调"""
        if fail_count > 0:
            self._log_message(f"轮询完成: {success_count}成功, {fail_count}失败", "WARNING")
        # 更新状态栏统计即可，不重建设备树

    @Slot()
    def _on_mcgsm_show_history(self) -> None:
        from PySide6.QtWidgets import QComboBox
        from PySide6.QtCore import Qt, QDateTime as _QDt
        from ui.widgets.history_chart_widget import HistoryChartWidget

        dialog = QDialog(self)
        dialog.setWindowTitle("MCGS历史数据查看")
        dialog.resize(1100, 800)

        layout = QVBoxLayout(dialog)

        # ── 顶部控制栏 ──
        control_bar = QHBoxLayout()

        control_bar.addWidget(QLabel("设备:"))
        device_combo = QComboBox()
        device_combo.setMinimumWidth(180)
        reader = None
        if self._mcgs_controller:
            reader = self._mcgs_controller.get_reader()
        device_ids = reader.list_devices() if reader else []
        for dev_id in device_ids:
            config = reader.get_device_config(dev_id) if reader else None
            display_name = getattr(config, "name", dev_id) or dev_id if config else dev_id
            label = f"{display_name} ({dev_id})"
            device_combo.addItem(label, dev_id)
        control_bar.addWidget(device_combo)

        control_bar.addWidget(QLabel("参数:"))
        param_combo = QComboBox()
        param_combo.setMinimumWidth(140)
        param_combo.addItem("全部参数", "__ALL__")
        control_bar.addWidget(param_combo)

        # 计算默认时间范围
        from datetime import datetime as _dt_local

        _now_ts = _dt_local.now()
        _today_9am = _now_ts.replace(hour=9, minute=0, second=0, microsecond=0)

        control_bar.addWidget(QLabel("起始:"))
        dt_start = QDateTimeEdit(calendarPopup=True)
        dt_start.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        dt_start.setDateTime(_QDt(_today_9am))
        dt_start.setFixedWidth(185)
        dt_start.setFixedHeight(30)
        dt_start.setStyleSheet(
            "QDateTimeEdit { padding: 2px 4px; border: 1px solid #c0c0c0; " "border-radius: 3px; background: white; }"
        )
        control_bar.addWidget(dt_start)

        control_bar.addWidget(QLabel("结束:"))
        dt_end = QDateTimeEdit(calendarPopup=True)
        dt_end.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        dt_end.setDateTime(_QDt.currentDateTime())
        dt_end.setFixedWidth(185)
        dt_end.setFixedHeight(30)
        dt_end.setStyleSheet(
            "QDateTimeEdit { padding: 2px 4px; border: 1px solid #c0c0c0; " "border-radius: 3px; background: white; }"
        )
        control_bar.addWidget(dt_end)

        # 快捷时间范围下拉框
        range_combo = QComboBox()
        range_combo.addItems(["最近1小时", "最近24小时", "最近1天", "最近1周"])
        range_combo.setCurrentIndex(0)  # 默认最近1小时
        range_combo.setToolTip("选择后自动应用快捷时间范围")
        control_bar.addWidget(range_combo)

        refresh_btn = QPushButton("加载")
        control_bar.addWidget(refresh_btn)

        # 合并导出按钮 - 点击弹出格式选择菜单
        export_btn = QPushButton("导出 ▼")

        def _show_export_menu():
            """点击导出按钮时创建菜单（延迟绑定，避免导出函数尚未定义）"""
            menu = QMenu(dialog)
            menu.addAction("CSV文件 (*.csv)", on_export_csv)
            menu.addAction("HTML报表 (*.html)", on_export_html)
            menu.addAction("Excel报表 (*.xlsx)", on_export_excel)
            menu.exec(export_btn.mapToGlobal(export_btn.rect().bottomLeft()))

        export_btn.clicked.connect(_show_export_menu)
        control_bar.addWidget(export_btn)

        control_bar.addStretch()
        layout.addLayout(control_bar)

        # ── 图表 + 表格 分割区 ──
        splitter = QSplitter(Qt.Orientation.Vertical)

        chart_widget = HistoryChartWidget("mcgsm_viewer")
        splitter.addWidget(chart_widget)

        data_table = QTableWidget()
        data_table.setAlternatingRowColors(True)
        data_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        data_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        data_table.horizontalHeader().setStretchLastSection(True)
        data_table.setMinimumHeight(150)
        splitter.addWidget(data_table)

        splitter.setStretchFactor(0, 3)  # 图表占 3/4
        splitter.setStretchFactor(1, 1)  # 表格占 1/4
        layout.addWidget(splitter, 1)

        # 存储当前加载的数据（用于导出）
        _current_data: Dict[str, List[Tuple[str, float]]] = {}  # {param: [(ts_str, val), ...]}
        _current_params: List[str] = []
        _current_dev_id: str = ""
        _current_hours: int = 0
        _pending_task: Optional[_HistoryQueryTask] = None

        def _get_param_display_map(dev_id: str) -> Dict[str, str]:
            """获取参数名→显示名的映射（含描述信息），格式：name（description）"""
            cfg = reader.get_device_config(dev_id) if reader else None
            pts = getattr(cfg, "points", []) or []
            display_map = {}
            for p in pts:
                p_name = getattr(p, "name", "")
                if not p_name:
                    continue
                p_desc = getattr(p, "description", "") or getattr(p, "unit", "")
                if p_desc:
                    display_map[p_name] = f"{p_name}（{p_desc}）"
                else:
                    display_map[p_name] = p_name
            return display_map

        def _collect_param_names(dev_id: str) -> List[str]:
            """获取设备所有的参数名称列表"""
            cfg = reader.get_device_config(dev_id) if reader else None
            pts = getattr(cfg, "points", []) or []
            return [getattr(p, "name", "") for p in pts if getattr(p, "name", "")]

        def _populate_param_combo(dev_id: str):
            """填充参数下拉框"""
            param_combo.blockSignals(True)
            param_combo.clear()
            param_combo.addItem("全部参数", "__ALL__")
            cfg = reader.get_device_config(dev_id) if reader else None
            pts = getattr(cfg, "points", []) or []
            for p in pts:
                p_name = getattr(p, "name", "")
                p_unit = getattr(p, "unit", "")
                p_addr = getattr(p, "addr", "")
                if p_name:
                    display = f"{p_name}"
                    if p_unit:
                        display += f" ({p_unit})"
                    display += f" @{p_addr}"
                    param_combo.addItem(display, p_name)
            param_combo.blockSignals(False)

        def _get_dt_range(dev_id: str):
            """获取设备数据在 DB 中的时间范围，用于限制 QDateTimeEdit"""
            try:
                hs = self._mcgs_controller._service._history_service if self._mcgs_controller else None
                if hs and hs.storage:
                    import sqlite3

                    cur = hs.storage._conn.cursor()
                    cur.execute("SELECT MIN(timestamp), MAX(timestamp) FROM mcgs_history WHERE device_id=?", (dev_id,))
                    row = cur.fetchone()
                    if row and row[0] and row[1]:
                        from datetime import datetime as _dtp

                        ts_min = _QDt.fromString(row[0].split(".")[0], "yyyy-MM-dd HH:mm:ss")
                        ts_max = _QDt.fromString(row[1].split(".")[0], "yyyy-MM-dd HH:mm:ss")
                        return ts_min, ts_max
            except Exception:
                pass
            return None, None

        def _update_dt_limits(dev_id: str):
            """根据设备数据的实际时间范围更新 QDateTimeEdit 上限（不下限，避免快捷时间被钳位）"""
            ts_min, ts_max = _get_dt_range(dev_id)
            if ts_min and ts_max:
                # 只限制最大值，不限制最小值
                # 如果设了 setMinimumDateTime，当用户选"最近1小时"时
                # 计算出的起始时间可能早于 ts_min，QDateTimeEdit 会自动钳位到 ts_min
                # 导致起始=结束，数据展示异常
                dt_start.setMaximumDateTime(ts_max)
                dt_end.setMaximumDateTime(ts_max)

        def _apply_shortcut(hours_delta: float, from_today_start: bool = False):
            """应用快捷时间选择"""
            from datetime import datetime as _dts

            now_py = _dts.now()
            if from_today_start:
                start_py = now_py.replace(hour=0, minute=0, second=0, microsecond=0)
            else:
                start_py = now_py - __import__("datetime").timedelta(hours=hours_delta)
            dt_start.setDateTime(_QDt(start_py))
            dt_end.setDateTime(_QDt.currentDateTime())

        def _populate_table(
            data_by_param: Dict[str, List[Tuple[str, float]]],
            param_names: List[str],
            display_names: Optional[List[str]] = None,
        ):
            """填充数据表格（宽表格式：时间戳|param1|param2|...）"""
            nonlocal _current_data
            _current_data.clear()
            _current_data.update(data_by_param)

            if not data_by_param:
                data_table.setRowCount(0)
                data_table.setColumnCount(0)
                return

            # 以第一个参数的行数为基准
            first_param = param_names[0] if param_names else list(data_by_param.keys())[0]
            rows = data_by_param.get(first_param, [])
            cols = len(param_names) + 1  # 时间戳 + 各参数

            # 表头使用显示名（含描述）
            headers = ["时间戳"] + ((display_names or param_names) if display_names else param_names)

            data_table.clear()
            data_table.setColumnCount(cols)
            data_table.setHorizontalHeaderLabels(headers)

            data_table.setRowCount(len(rows))
            for ri, (ts_str, _) in enumerate(rows):
                # 时间戳居中
                ts_item = QTableWidgetItem(ts_str)
                ts_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                data_table.setItem(ri, 0, ts_item)
                for ci, pn in enumerate(param_names):
                    param_data = data_by_param.get(pn, [])
                    if ri < len(param_data):
                        _, val = param_data[ri]
                        item = QTableWidgetItem(f"{val:.2f}")
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        data_table.setItem(ri, ci + 1, item)

            # 所有列均匀拉伸，每个格子大小相同
            for col_i in range(cols):
                data_table.horizontalHeader().setSectionResizeMode(col_i, QHeaderView.ResizeMode.Stretch)
            # 统一行高
            data_table.verticalHeader().setDefaultSectionSize(30)

        def load_chart_data():
            dev_id = device_combo.currentData()
            if not dev_id:
                return

            dt_start_py = dt_start.dateTime().toPython()
            dt_end_py = dt_end.dateTime().toPython()

            if dt_start_py >= dt_end_py:
                QMessageBox.warning(dialog, "时间错误", "起始时间必须早于结束时间")
                return

            nonlocal _current_dev_id, _pending_task
            _current_dev_id = dev_id

            # 取消之前未完成的任务
            if _pending_task is not None:
                _pending_task.cancel()
                _pending_task = None

            # 按钮状态 + 任务栏标题
            refresh_btn.setText("加载中...")
            refresh_btn.setEnabled(False)
            dialog.setWindowTitle("MCGS历史数据查看 - 刷新中...")

            if self._mcgs_controller and self._mcgs_controller.is_history_available():
                _load_from_history_service(dev_id, dt_start_py, dt_end_py)
            else:
                _load_from_memory(dev_id)

        def _set_loading_done():
            """恢复按钮状态和窗口标题"""
            refresh_btn.setText("刷新")
            refresh_btn.setEnabled(True)
            dialog.setWindowTitle("MCGS历史数据查看")

        def _load_from_history_service(dev_id, start_time, end_time):
            # 确定参数列表（支持全部参数/单参数选择）
            param_key = param_combo.currentData()
            if param_key and param_key != "__ALL__":
                param_names = [param_key]
            else:
                param_names = _collect_param_names(dev_id)

            nonlocal _current_params, _current_hours
            _current_params = param_names
            _current_hours = 0

            if not param_names:
                data_table.setRowCount(0)
                data_table.setColumnCount(0)
                QMessageBox.information(dialog, "无数据", f"设备 [{dev_id}] 没有配置参数")
                _set_loading_done()
                return

            # 移除之前残留的加载指示器
            for _i in range(splitter.count()):
                _w = splitter.widget(_i)
                if isinstance(_w, (QLabel, QProgressBar)):
                    _w.deleteLater()

            # 使用 QProgressDialog 模态进度对话框，不影响布局
            progress_dialog = QProgressDialog("正在加载历史数据...", "", 0, len(param_names), dialog)
            progress_dialog.setWindowTitle("加载中")
            progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
            progress_dialog.setCancelButton(None)
            progress_dialog.setAutoClose(True)
            progress_dialog.setMinimumDuration(0)
            progress_dialog.setStyleSheet(
                "QProgressDialog { min-width: 360px; min-height: 120px; }"
                "QProgressBar { border: 1px solid #B0C4DE; border-radius: 6px; "
                "background: #F0F4FF; text-align: center; font-size: 13px; font-weight: bold; "
                "color: #1A56DB; height: 24px; }"
                "QProgressBar::chunk { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, "
                "stop:0 #3B82F6, stop:1 #1A56DB); border-radius: 5px; }"
                "QLabel { font-size: 13px; color: #333; }"
            )
            progress_dialog.show()

            def _on_result(result_data):
                """结果回调（在主线程）"""
                data_by_param, param_names_out, has_data = result_data
                nonlocal _pending_task
                _pending_task = None
                _set_loading_done()

                # 对话框关闭保护
                if not dialog.isVisible():
                    return

                # 关闭进度对话框
                progress_dialog.close()
                progress_dialog.deleteLater()

                if has_data:
                    # 计算参数显示名（含描述）
                    _display_map = _get_param_display_map(dev_id)
                    _display_names = [_display_map.get(n, n) for n in param_names_out]
                    _dev_label = device_combo.currentText()
                    # 降采样：减少图表渲染压力
                    for pname in data_by_param:
                        total_before = len(data_by_param[pname])
                        data_by_param[pname] = _downsample_data(data_by_param[pname], 5000)
                        total_after = len(data_by_param[pname])
                        if total_before != total_after:
                            logger.info("降采样 [%s]: %d → %d 点", pname, total_before, total_after)

                    # 将字符串时间戳转换为 datetime 对象
                    _dt_data = {}
                    for pname, points in data_by_param.items():
                        from datetime import datetime as _dt

                        _dt_data[pname] = []
                        for ts_str, val in points:
                            try:
                                dt = _dt.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
                            except (ValueError, TypeError):
                                dt = _dt.now()
                            _dt_data[pname].append((dt, val))

                    logger.info("图表加载: %d 参数, 总 %d 点", len(_dt_data), sum(len(v) for v in _dt_data.values()))

                    chart_widget.set_bulk_data(
                        _dt_data, param_names_out, display_names=_display_names, device_display_name=_dev_label
                    )
                    from PySide6.QtCore import QCoreApplication

                    QCoreApplication.processEvents()
                    chart_widget.refresh_chart()
                    chart_widget.repaint()
                    _populate_table(data_by_param, param_names_out, display_names=_display_names)
                else:
                    data_table.setRowCount(0)
                    data_table.setColumnCount(0)
                    QMessageBox.information(dialog, "无数据", f"未找到设备 [{dev_id}] 的历史数据")

            def _on_progress(progress_data):
                """进度更新回调（在主线程）"""
                if not dialog.isVisible():
                    return
                current, total = progress_data
                progress_dialog.setValue(current)
                progress_dialog.setLabelText(f"正在加载... ({current}/{total})")

            def _on_error(error_msg):
                """错误回调（在主线程）"""
                nonlocal _pending_task
                _pending_task = None
                _set_loading_done()

                if not dialog.isVisible():
                    return

                progress_dialog.close()
                progress_dialog.deleteLater()
                logger.error("历史数据加载失败: %s", error_msg)
                QMessageBox.warning(dialog, "查询失败", f"历史数据查询异常:\n{error_msg}")

            # 获取 HistoryService 实例
            hs = (
                self._mcgs_controller._service._history_service
                if (
                    self._mcgs_controller
                    and self._mcgs_controller._service
                    and self._mcgs_controller._service._history_service
                )
                else None
            )

            if hs is None:
                progress_dialog.close()
                progress_dialog.deleteLater()
                _set_loading_done()
                QMessageBox.warning(dialog, "服务不可用", "历史数据服务未初始化")
                return

            # 使用 QThreadPool + QRunnable 后台查询
            task = _HistoryQueryTask(hs, dev_id, param_names, hours=24, start_time=start_time, end_time=end_time)
            task.signals.progress.connect(_on_progress)
            task.signals.result.connect(_on_result)
            task.signals.error.connect(_on_error)
            QThreadPool.globalInstance().start(task)
            _pending_task = task

        def _load_from_memory(dev_id):
            """从内存缓存加载历史数据（降级路径）"""
            card_history = (
                self._monitor_controller._card_history if hasattr(self._monitor_controller, "_card_history") else {}
            )
            if not card_history:
                data_table.setRowCount(0)
                data_table.setColumnCount(0)
                QMessageBox.information(dialog, "无数据", "暂无历史数据（设备需要先连接并采集数据）")
                _set_loading_done()
                return

            # 支持单参数选择
            param_key = param_combo.currentData()
            if param_key and param_key != "__ALL__":
                params_to_show = [param_key] if param_key in card_history else []
            else:
                params_to_show = list(card_history.keys())

            nonlocal _current_params
            _current_params = params_to_show

            # 在内存模式下，先构建批量数据再一次性更新UI
            has_data = False
            data_by_param: Dict[str, List[Tuple[str, float]]] = {}
            chart_data: Dict[str, List[Tuple[datetime, float]]] = {}
            import time as _time
            from datetime import datetime as _dt

            for pname in params_to_show:
                history = card_history.get(pname, [])
                if history:
                    has_data = True
                    points_list = []
                    dt_points = []
                    ts_base = _time.time() - len(history)
                    for i, val in enumerate(history):
                        try:
                            dt = _dt.fromtimestamp(ts_base + i)
                            ts_str = dt.strftime("%Y-%m-%d %H:%M:%S")
                            points_list.append((ts_str, float(val)))
                            dt_points.append((dt, float(val)))
                        except (ValueError, TypeError):
                            pass
                    data_by_param[pname] = points_list
                    chart_data[pname] = dt_points

            if has_data:
                _display_map = _get_param_display_map(dev_id)
                _display_names = [_display_map.get(n, n) for n in params_to_show]
                _dev_label = device_combo.currentText()
                chart_widget.set_bulk_data(
                    chart_data, params_to_show, display_names=_display_names, device_display_name=_dev_label
                )
                _populate_table(data_by_param, params_to_show, display_names=_display_names)
            else:
                data_table.setRowCount(0)
                data_table.setColumnCount(0)
                QMessageBox.information(
                    dialog, "无数据", f"未找到 [{dev_id}] 的内存历史数据（请先在监控页面查看实时数据）"
                )
            _set_loading_done()

        def _export_html(file_path: str):
            """导出为自包含 HTML 文件（含数据表格+内嵌Base64 Chart.js曲线图，可直接浏览器打开）"""
            if not _current_data:
                QMessageBox.warning(dialog, "导出失败", "没有数据可供导出，请先刷新加载数据")
                return

            # 计算参数显示名（含描述）
            _html_disp_map = _get_param_display_map(_current_dev_id) if _current_dev_id else {}
            _html_disp_names = [_html_disp_map.get(pn, pn) for pn in _current_params]

            # 准备图表数据
            chart_colors = [
                "#3B82F6",
                "#EF4444",
                "#10B981",
                "#F59E0B",
                "#8B5CF6",
                "#EC4899",
                "#06B6D4",
                "#84CC16",
                "#6366F1",
                "#F97316",
            ]

            # 生成数据数组（用JSON安全格式）
            import json as _json

            datasets_js = []
            for ci, pn in enumerate(_current_params):
                param_data = _current_data.get(pn, [])
                vals = [v for _, v in param_data]
                color = chart_colors[ci % len(chart_colors)]
                _label = _html_disp_names[ci] if ci < len(_html_disp_names) else pn
                datasets_js.append(
                    _json.dumps(
                        {
                            "label": _label,
                            "data": vals,
                            "borderColor": color,
                            "backgroundColor": color + "40",
                            "borderWidth": 1.5,
                            "tension": 0.1,
                            "pointRadius": 0.5,
                        },
                        ensure_ascii=False,
                    )
                )

            first_param = _current_params[0] if _current_params else list(_current_data.keys())[0]
            time_labels = _current_data.get(first_param, [])
            ts_labels = [t for t, _ in time_labels]
            labels_json = _json.dumps(ts_labels, ensure_ascii=False)

            lines = []
            lines.append("<!DOCTYPE html><html><head><meta charset='utf-8'>")
            lines.append("<title>历史数据报表</title>")
            # 内嵌 Chart.js 4.x 精简版（避免外网加载失败）
            lines.append("<script>")
            lines.append(
                "var Chart=function(){return{Chart:class{constructor(e,t){this.ctx=e;"
                "this.config=t;this.draw()}"
                "draw(){const e=this.ctx,t=this.config;"
                "e.clearRect(0,0,e.canvas.width,e.canvas.height);"
                "const n=t.data;if(!n||!n.labels||!n.datasets)return;"
                "const r=e.canvas.parentElement;"
                "e.canvas.width=r.clientWidth||800;"
                "e.canvas.height=r.clientHeight||350;"
                "const i=n.datasets,o=n.labels;"
                "let a=1/0,s=-1/0;i.forEach(e=>{e.data.forEach(e=>{e<a&&(a=e),"
                "e>s&&(s=e)})});a===1/0&&(a=0,s=100);const l=s-a||1;"
                "const c={top:30,bottom:40,left:50,right:20};"
                "const u=e.canvas.width-c.left-c.right;"
                "const d=e.canvas.height-c.top-c.bottom;"
                "e.strokeStyle='#e5e7eb';e.lineWidth=0.5;"
                "for(let t=0;t<=5;t++){const n=c.top+d*t/5;"
                "e.beginPath();e.moveTo(c.left,n);e.lineTo(c.left+u,n);e.stroke();"
                "e.fillStyle='#6b7280';e.font='10px sans-serif';"
                "e.textAlign='right';e.fillText((s-(s-a)*t/5).toFixed(1),c.left-5,n+3)}"
                "const p=o.length;const h=u/(p-1||1);"
                "i.forEach((n,r)=>{const i=n.data;const o=n.borderColor||'#333';"
                "e.strokeStyle=o;e.lineWidth=n.borderWidth||1.5;e.beginPath();"
                "i.forEach((n,r)=>{const i=c.left+r*h;"
                "const f=c.top+d-(n-a)/l*d;"
                "r===0?e.moveTo(i,f):e.lineTo(i,f)});e.stroke();"
                "e.fillStyle=o;e.font='10px sans-serif';e.textAlign='center';"
                "e.fillText(n.label||'',c.left+u/2,c.top-10)});"
                "e.fillStyle='#6b7280';e.font='10px sans-serif';"
                "e.textAlign='center';"
                "const g=Math.max(1,Math.floor(p/10));"
                "o.forEach((t,n)=>{n%g===0&&e.fillText(t,c.left+n*h,c.top+d+15)})}}}}}();"
            )
            lines.append("</script>")
            lines.append("<style>")
            lines.append("*{box-sizing:border-box}")
            lines.append("body{font-family:'Microsoft YaHei',sans-serif;margin:20px;color:#333}")
            lines.append("h1{color:#1a56db;font-size:20px;margin-bottom:5px}")
            lines.append("h2{color:#6b7280;font-size:13px;font-weight:400;margin-bottom:15px}")
            lines.append(
                "#chartBox{width:100%;height:380px;border:1px solid #e5e7eb;"
                "border-radius:6px;margin-bottom:20px;position:relative}"
            )
            lines.append("#chartBox canvas{width:100%;height:100%}")
            lines.append(
                "#chartBox .loading{position:absolute;top:50%;left:50%;"
                "transform:translate(-50%,-50%);color:#9ca3af;font-size:14px}"
            )
            lines.append("table{border-collapse:collapse;width:100%;font-size:11px}")
            lines.append(
                "th{background:#1a56db;color:#fff;padding:6px 8px;text-align:center;"
                "border:1px solid #1a56db;position:sticky;top:0}"
            )
            lines.append("td{padding:4px 8px;border:1px solid #e5e7eb;text-align:right}")
            lines.append("td:first-child{text-align:center;white-space:nowrap}")
            lines.append("tr:nth-child(even){background:#f9fafb}")
            lines.append("tr:hover{background:#eff6ff}")
            lines.append(".table-wrap{max-height:450px;overflow-y:auto;" "border:1px solid #e5e7eb;border-radius:6px}")
            lines.append("</style></head><body>")

            lines.append(f"<h1>历史数据报表</h1>")
            lines.append(
                f"<h2>设备: {device_combo.currentText()} &nbsp;|&nbsp; "
                f"时间段: 最近{_current_hours}小时 &nbsp;|&nbsp; "
                f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} &nbsp;|&nbsp; "
                f"数据点: {len(time_labels)}条/参数</h2>"
            )

            # ── 内嵌 Canvas 曲线图 ──
            lines.append(
                "<div id='chartBox'><span class='loading'>正在渲染曲线图...</span>"
                "<canvas id='historyChart'></canvas></div>"
            )
            lines.append("<script>")
            lines.append("(function(){")
            lines.append("var c=document.getElementById('historyChart');")
            lines.append("var p=c.parentElement;")
            lines.append(f"var labels={labels_json};")
            lines.append(f"var datasets={','.join('['+ds+']' for ds in datasets_js)};")
            # 将datasets_js合并为数组
            lines.append("var allData={labels:labels,datasets:[" + ",".join(datasets_js) + "]};")
            lines.append(
                "try{new Chart(c.getContext('2d'),{type:'line',data:allData})"
                "}catch(e){p.innerHTML='<div style=\"padding:40px;text-align:center;"
                "color:#ef4444\">曲线图渲染失败，请稍后重试或使用Excel导出</div>'}"
            )
            lines.append("})();")
            lines.append("</script>")

            # ── 数据表格 ──
            lines.append("<div class='table-wrap'>")
            lines.append("<table><thead><tr>")
            lines.append("<th>时间戳</th>")
            for pn in _html_disp_names:
                lines.append(f"<th>{pn}</th>")
            lines.append("</tr></thead><tbody>")

            rows = time_labels
            for ri, (ts_str, _) in enumerate(rows):
                lines.append("<tr>")
                lines.append(f"<td>{ts_str}</td>")
                for pn in _current_params:
                    param_data = _current_data.get(pn, [])
                    if ri < len(param_data):
                        _, val = param_data[ri]
                        lines.append(f"<td>{val:.4f}</td>")
                    else:
                        lines.append("<td></td>")
                lines.append("</tr>")

            lines.append("</tbody></table></div>")
            lines.append("</body></html>")

            html = "\n".join(lines)
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(html)
            return True

        def on_export_csv():
            from pathlib import Path

            file_path, _ = QFileDialog.getSaveFileName(
                dialog, "导出CSV", "history_data.csv", "CSV Files (*.csv);;长表格式CSV (*.csv)"
            )
            if not file_path:
                return

            dialog.setWindowTitle("MCGS历史数据查看 - 导出CSV文件中...")
            # 先用当前内存数据（如果已有加载）
            if _current_data and _current_params:
                # 导出内存中已加载的数据
                import csv as _csv

                try:
                    # 计算参数显示名（含描述）
                    _disp_map_csv = _get_param_display_map(_current_dev_id) if _current_dev_id else {}
                    _csv_headers = ["时间戳"] + [_disp_map_csv.get(pn, pn) for pn in _current_params]
                    with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                        writer = _csv.writer(f)
                        writer.writerow(_csv_headers)
                        first_param = _current_params[0]
                        rows = _current_data.get(first_param, [])
                        for ri, (ts_str, _) in enumerate(rows):
                            row = [ts_str]
                            for pn in _current_params:
                                param_data = _current_data.get(pn, [])
                                if ri < len(param_data):
                                    _, val = param_data[ri]
                                    row.append(f"{val:.4f}")
                                else:
                                    row.append("")
                            writer.writerow(row)
                    QMessageBox.information(dialog, "导出成功", f"数据已保存至:\n{file_path}")
                    dialog.setWindowTitle("MCGS历史数据查看")
                    return
                except Exception as e:
                    logger.error("CSV导出失败: %s", e)
                    QMessageBox.warning(dialog, "导出失败", f"CSV导出失败: {e}")
                dialog.setWindowTitle("MCGS历史数据查看")
                return

            # 回退到 HistoryService 导出
            dev_id = device_combo.currentData()
            if not dev_id:
                dialog.setWindowTitle("MCGS历史数据查看")
                return

            if self._mcgs_controller and self._mcgs_controller.is_history_available():
                param_names = _collect_param_names(dev_id)
                if not param_names:
                    QMessageBox.information(dialog, "无数据", f"设备 [{dev_id}] 没有配置参数")
                    dialog.setWindowTitle("MCGS历史数据查看")
                    return
                success = self._mcgs_controller.export_history_csv(file_path, dev_id, param_names)
                if success:
                    QMessageBox.information(dialog, "导出成功", f"数据已保存至:\n{file_path}")
                else:
                    QMessageBox.warning(dialog, "导出失败", "导出历史数据失败")
            else:
                QMessageBox.warning(dialog, "导出失败", "历史服务不可用")
            dialog.setWindowTitle("MCGS历史数据查看")

        def on_export_html():
            if not _current_data:
                QMessageBox.warning(dialog, "无数据", "请先刷新加载数据后再导出")
                return
            file_path, _ = QFileDialog.getSaveFileName(
                dialog, "导出HTML报表", "history_report.html", "HTML Files (*.html)"
            )
            if file_path:
                dialog.setWindowTitle("MCGS历史数据查看 - 导出HTML报表中...")
                try:
                    _export_html(file_path)
                    QMessageBox.information(dialog, "导出成功", f"HTML报表已保存至:\n{file_path}")
                except Exception as e:
                    logger.error("HTML导出失败: %s", e)
                    QMessageBox.warning(dialog, "导出失败", f"HTML导出失败: {e}")
                dialog.setWindowTitle("MCGS历史数据查看")

        def _export_excel(file_path: str):
            """导出为Excel文件（仅数据表格，不含曲线图）"""
            if not _current_data:
                return False

            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()

            # ── 唯一 Sheet: 数据表 ──
            ws_data = wb.active
            ws_data.title = "数据表"

            # 表头样式
            header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=10)
            header_fill = PatternFill(start_color="1A56DB", end_color="1A56DB", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin", color="D1D5DB"),
                right=Side(style="thin", color="D1D5DB"),
                top=Side(style="thin", color="D1D5DB"),
                bottom=Side(style="thin", color="D1D5DB"),
            )

            # 写入表头（使用显示名）
            _excel_disp_map = _get_param_display_map(_current_dev_id) if _current_dev_id else {}
            excel_headers = ["时间戳"] + [_excel_disp_map.get(pn, pn) for pn in _current_params]
            for ci, h in enumerate(excel_headers, 1):
                cell = ws_data.cell(row=1, column=ci, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # 写入数据
            first_param = _current_params[0]
            rows = _current_data.get(first_param, [])
            alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            data_font = Font(name="Microsoft YaHei", size=10)
            from datetime import datetime as _dt_excel

            def _parse_ts_safe(ts_str):
                try:
                    return _dt_excel.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    try:
                        return _dt_excel.strptime(ts_str, "%Y-%m-%d %H:%M:%S.%f")
                    except ValueError:
                        return ts_str

            for ri, (ts_str, _) in enumerate(rows):
                row_num = ri + 2
                ts_dt = _parse_ts_safe(ts_str)
                cell = ws_data.cell(row=row_num, column=1, value=ts_dt)
                cell.font = data_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
                cell.number_format = "YYYY-MM-DD HH:MM:SS"
                if ri % 2 == 1:
                    cell.fill = alt_fill
                for ci, pn in enumerate(_current_params):
                    param_data = _current_data.get(pn, [])
                    val = param_data[ri][1] if ri < len(param_data) else None
                    cell = ws_data.cell(row=row_num, column=ci + 2, value=val)
                    cell.font = data_font
                    cell.number_format = "0.0000"
                    cell.alignment = Alignment(horizontal="right")
                    cell.border = thin_border
                    if ri % 2 == 1:
                        cell.fill = alt_fill

            # 自动列宽
            for ci, h in enumerate(excel_headers, 1):
                col_letter = get_column_letter(ci)
                max_len = max(
                    len(str(ws_data.cell(row=r, column=ci).value or "")) for r in range(1, min(len(rows) + 2, 200))
                )
                ws_data.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 30)

            wb.save(file_path)
            return True

        def on_export_excel():
            if not _current_data:
                QMessageBox.warning(dialog, "无数据", "请先刷新加载数据后再导出")
                return
            file_path, _ = QFileDialog.getSaveFileName(
                dialog, "导出Excel报表", "history_report.xlsx", "Excel Files (*.xlsx)"
            )
            if file_path:
                dialog.setWindowTitle("MCGS历史数据查看 - 导出Excel报表中...")
                try:
                    _export_excel(file_path)
                    QMessageBox.information(dialog, "导出成功", f"Excel报表已保存至:\n{file_path}")
                except Exception as e:
                    logger.error("Excel导出失败: %s", e)
                    QMessageBox.warning(dialog, "导出失败", f"Excel导出失败: {e}")
                dialog.setWindowTitle("MCGS历史数据查看")

        # 信号连接
        def _on_device_changed():
            dev_id = device_combo.currentData()
            if dev_id:
                _populate_param_combo(dev_id)
                _update_dt_limits(dev_id)
                # 切换设备时重置时间范围（最近1小时），避免 dt_start >= dt_end
                _apply_shortcut(1)
                load_chart_data()

        def _on_range_changed():
            """快捷时间范围下拉选择时自动应用并加载数据"""
            text = range_combo.currentText()
            if text == "最近1小时":
                _apply_shortcut(1)
            elif text == "最近24小时":
                _apply_shortcut(24)
            elif text == "最近1天":
                _apply_shortcut(24, from_today_start=True)
            elif text == "最近1周":
                _apply_shortcut(168)
            load_chart_data()

        device_combo.currentIndexChanged.connect(_on_device_changed)
        range_combo.currentIndexChanged.connect(_on_range_changed)
        param_combo.currentIndexChanged.connect(load_chart_data)
        refresh_btn.clicked.connect(load_chart_data)

        if device_ids:
            _populate_param_combo(device_ids[0])
            _update_dt_limits(device_ids[0])
            # 默认应用最近1小时
            _apply_shortcut(1)
            load_chart_data()

        dialog.exec()

    @Slot()
    def _on_show_mcgsm_config(self) -> None:
        """
        打开MCGS设备可视化配置编辑器（菜单/工具栏按钮触发）

        功能：
        - 加载当前 devices.json 配置
        - 提供可视化编辑界面（4个Tab）
        - 保存后自动刷新读取器
        """
        focus_id = self._current_device_id or None
        self._open_mcgsm_config_dialog(focus_device_id=focus_id)

    def _open_mcgsm_config_dialog(self, focus_device_id: str = None) -> bool:
        """
        打开MCGS配置对话框（内部调用版本）

        Args:
            focus_device_id: 可选，指定要聚焦的设备ID（用于连接失败时的引导）

        Returns:
            bool: 用户是否保存了配置
        """
        from pathlib import Path

        config_path = Path(__file__).parent.parent / "config" / "devices.json"

        # 创建对话框实例
        dialog = MCGSConfigDialog(config_path, parent=self)

        # 如果指定了设备ID，尝试切换到该设备的配置
        if focus_device_id:
            dialog.setFocusDevice(focus_device_id)

        # 以模态方式打开对话框
        result = dialog.exec()

        if result == QDialog.DialogCode.Accepted:
            logger.info(f"MCGS配置已通过编辑器更新: {config_path}")

            if self._mcgs_controller:
                self._mcgs_controller.reset_reader()

            self._refresh_device_list(self._search_edit.text())
            self._update_status_bar()
            self._status_msg_label.setText("MCGS配置已更新，设备列表已刷新")

            # 强制刷新当前设备的数据卡片（使描述/单位即时生效）
            if self._current_device_id:
                QTimer.singleShot(500, lambda: self._refresh_cards_for_device(self._current_device_id))

            return True
        else:
            # 用户取消了对话框
            logger.info("MCGS配置编辑器已取消")
            return False

    # ══════════════════════════════════════════════════════════
    # ✅ 新增：操作撤销相关方法
    # ══════════════════════════════════════════════════════════

    def _on_undo_executed(self, req_id: str) -> None:
        """
        撤销操作执行槽函数

        当 OperationUndoManager 发射 undo_executed 信号时调用，
        通过 WriteOperationManager 发起安全恢复写入。

        Args:
            req_id: 原操作的请求ID
        """
        if not hasattr(self, "_undo_mgr"):
            return

        record = self._undo_mgr.get_record_by_req_id(req_id)
        if record is None:
            logger.error("未找到撤销记录: %s", req_id)
            return

        logger.info(
            "执行撤销操作 [参数=%s, 恢复为=%s]",
            record.param_name,
            DeviceController.format_undo_value(record.previous_value),
        )

        device_id = record.device_id
        if device_id != self._current_device_id:
            logger.info("切换到设备 %s 以执行撤销", device_id)

        try:
            device = self._device_manager.get_device(device_id)
            if not device:
                self._undo_mgr.mark_undo_failed(req_id, "设备对象不存在")
                return

            rp = None
            if hasattr(device, "register_points"):
                rp = device.get_register_point_by_name(record.param_name)

            if rp is None:
                self._undo_mgr.mark_undo_failed(req_id, f"未找到参数配置: {record.param_name}")
                return

            restore_value = (
                bool(record.previous_value) if record.operation_type == "coil_write" else record.previous_value
            )

            undo_req_id = self._write_manager.request_write(
                device_id=device_id,
                param_name=record.param_name,
                value=restore_value,
                config=rp,
                skip_confirm=True,
            )
            self._undo_req_map[undo_req_id] = req_id

        except Exception as e:
            error_msg = f"撤销异常: {str(e)}"
            logger.error(error_msg)
            self._undo_mgr.mark_undo_failed(req_id, error_msg)
            self._status_msg_label.setText(error_msg)

    def _on_undo_failed(self, req_id: str, reason: str) -> None:
        """
        撤销失败槽函数

        Args:
            req_id: 请求ID
            reason: 失败原因
        """
        logger.error("撤销操作失败 [req=%s]: %s", req_id, reason)
        self._status_msg_label.setText(f"撤销失败: {reason}")

    def undo_last_operation(self) -> None:
        """
        撤销最后一次操作（公开方法）

        可由工具栏按钮、快捷键Ctrl+Z或右键菜单调用。
        """
        if not hasattr(self, "_undo_mgr") or not self._undo_mgr.can_undo:
            from PySide6.QtWidgets import QMessageBox

            QMessageBox.information(self, "无法撤销", "没有可撤销的操作。")
            return

        # 权限检查
        if not self._permission_mgr.can_write:
            from PySide6.QtWidgets import QMessageBox

            QMessageBox.warning(self, "权限不足", "当前用户无权执行撤销操作。")
            return

        # 确认对话框
        from PySide6.QtWidgets import QMessageBox

        history = self._undo_mgr.get_undo_history(limit=1)
        if history:
            last_record = history[0]
            summary = last_record.display_summary

            reply = QMessageBox.question(
                self,
                "确认撤销",
                f"确定要撤销以下操作吗？\n\n"
                f"{summary}\n\n"
                f"这将把 '{last_record.param_name}' 恢复为 "
                f"{DeviceController.format_undo_value(last_record.previous_value)}",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )

            if reply == QMessageBox.StandardButton.Yes:
                self._undo_mgr.undo_last_operation()

    def show_undo_history(self) -> None:
        """显示撤销历史面板"""
        if not hasattr(self, "_undo_mgr"):
            return

        history = self._undo_mgr.get_undo_history(limit=10)

        if not history:
            from PySide6.QtWidgets import QMessageBox

            QMessageBox.information(self, "撤销历史", "暂无撤销记录。")
            return

        # 构建历史信息文本
        lines = ["最近10次操作历史：\n"]
        for i, record in enumerate(history, 1):
            status = "可撤销" if record.can_undo and not record.undone_at else "已撤销"
            time_str = record.executed_at.strftime("%H:%M:%S")
            lines.append(f"{i}. [{time_str}] {record.display_summary} ({status})")

        from PySide6.QtWidgets import QMessageBox

        QMessageBox.information(self, "撤销历史", "\n".join(lines))

    def closeEvent(self, event) -> None:
        self._settings.setValue("window/geometry", self.saveGeometry())
        self.cleanup()
        event.accept()

    # ══════════════════════════════════════════════════════════
    # v3.2 新增：写操作管理相关方法
    # ══════════════════════════════════════════════════════════

    def _show_write_confirmation(self, device_id: str, param_name: str, value: object, config: object) -> None:
        """
        显示写操作确认对话框（由 WriteOperationManager.confirm_required 信号触发）

        弹出 QMessageBox 询问用户是否确认写入操作。
        用户响应后调用 on_user_confirmed()。

        Args:
            device_id: 设备ID
            param_name: 参数名称
            value: 目标值
            config: RegisterPointConfig 实例
        """
        # 格式化显示值
        display_value = "ON (闭合)" if value else "OFF (断开)"

        # 构建确认消息
        message = (
            f"确定要执行以下写操作吗？\n\n"
            f"设备: {device_id}\n"
            f"参数: {param_name}\n"
            f"目标值: {display_value}\n\n"
            f"此操作将立即影响设备运行状态！"
        )

        reply = QMessageBox.question(
            self,
            "写操作确认",
            message,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )

        # 调用 WriteOperationManager 的用户确认回调
        if hasattr(self, "_pending_write_req_id") and self._pending_write_req_id:
            self._write_manager.on_user_confirmed(
                self._pending_write_req_id, approved=(reply == QMessageBox.StandardButton.Yes)
            )
            self._pending_write_req_id = None

    def _show_history_dialog(self):
        """显示历史数据对话框"""
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QLabel, QPushButton, QApplication
        from PySide6.QtCore import QThreadPool
        from ui.widgets.history_chart_widget import HistoryChartWidget

        current_device_id = getattr(self, "_current_device_id", None)
        if not current_device_id:
            QMessageBox.information(self, "提示", "请先选择一个设备")
            return

        hs = (
            self._mcgs_controller._service._history_service
            if (
                self._mcgs_controller
                and self._mcgs_controller._service
                and self._mcgs_controller._service._history_service
            )
            else None
        )

        dialog = QDialog(self)
        dialog.setWindowTitle(f"历史数据 - {current_device_id}")
        dialog.setMinimumSize(900, 600)
        layout = QVBoxLayout(dialog)

        history_widget = HistoryChartWidget()
        layout.addWidget(history_widget)

        status_label = QLabel("正在加载历史数据...")
        status_label.setStyleSheet("color: #666; font-size: 13px; padding: 8px;")
        layout.addWidget(status_label)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(dialog.accept)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)

        if hs is None:
            status_label.setText("历史数据服务未初始化")
            status_label.setStyleSheet("color: #E53E3E; font-size: 13px; padding: 8px;")
            dialog.exec()
            return

        from datetime import datetime, timedelta

        end_time = datetime.now()
        start_time = end_time - timedelta(hours=24)

        class _SimpleHistoryQueryTask(QRunnable):
            class Signals(QObject):
                result = Signal(object)
                error = Signal(str)

            def __init__(self, hs, dev_id, start, end):
                super().__init__()
                self._hs = hs
                self._dev_id = dev_id
                self._start = start
                self._end = end
                self.signals = self.Signals()

            def run(self):
                try:
                    result = self._hs.query_device_data(self._dev_id, self._start, self._end)
                    self.signals.result.emit(result)
                except Exception as e:
                    self.signals.error.emit(str(e))

        def _on_result(result):
            if result and isinstance(result, dict) and result:
                total_points = sum(len(v) for v in result.values())
                history_widget.set_bulk_data(
                    result,
                    list(result.keys()),
                    device_display_name=current_device_id,
                )
                status_label.setText(f"已加载 {total_points} 条历史数据")
                status_label.setStyleSheet("color: #38A169; font-size: 13px; padding: 8px;")
            else:
                status_label.setText("该设备暂无历史数据记录")
                status_label.setStyleSheet("color: #D69E2E; font-size: 13px; padding: 8px;")

        def _on_error(error_msg):
            status_label.setText(f"查询失败: {error_msg}")
            status_label.setStyleSheet("color: #E53E3E; font-size: 13px; padding: 8px;")
            logger.error("历史数据查询失败 [%s]: %s", current_device_id, error_msg)

        task = _SimpleHistoryQueryTask(hs, current_device_id, start_time, end_time)
        task.signals.result.connect(_on_result)
        task.signals.error.connect(_on_error)
        QThreadPool.globalInstance().start(task)

        dialog.exec()

    def _on_coil_write_request(self, param_name: str, value: bool) -> None:
        """
        处理线圈写请求（由 DynamicMonitorPanel.coil_write_requested 信号触发）

        这是用户点击可写线圈按钮时的入口点。
        通过 WriteOperationManager 发起安全写流程。

        Args:
            param_name: 参数名称
            value: 目标状态
        """
        if not self._current_device_id:
            QMessageBox.warning(self, "操作失败", "请先选择一个设备")
            return

        device = self._device_manager.get_device(self._current_device_id)
        if not device:
            QMessageBox.warning(self, "操作失败", "设备对象不存在")
            return

        # 查找配置
        if hasattr(device, "register_points"):
            rp = device.get_register_point_by_name(param_name)
        else:
            rp = None

        if rp is None:
            logger.warning("未找到参数配置: %s", param_name)
            return

        try:
            # 通过 WriteOperationManager 发起请求
            req_id = self._write_manager.request_write(
                device_id=self._current_device_id,
                param_name=param_name,
                value=value,
                config=rp,
            )

            # 保存当前请求ID（用于回调时使用）
            self._pending_write_req_id = req_id

            logger.info("线圈写请求已提交 [设备=%s, 参数=%s, req=%s]", self._current_device_id, param_name, req_id)

        except Exception as e:
            QMessageBox.critical(self, "操作失败", f"发起写请求失败:\n{str(e)}")
            logger.error("发起写请求失败: %s", str(e), exc_info=True)

    def _on_write_operation_result(self, req_id: str, operation_dict: dict) -> None:
        """
        处理写操作完成通知（由 WriteOperationManager.operation_result 信号触发）

        当用户确认/取消或执行完成后，此方法被调用。
        对于已确认的操作，需要实际调用 DeviceConnection 执行写入。

        Args:
            req_id: 请求ID
            operation_dict: 操作详情字典
        """
        status = operation_dict.get("status")
        device_id = operation_dict.get("device_id", "")
        param_name = operation_dict.get("param_name", "")

        logger.debug("写操作结果 [req=%s, 状态=%s, 设备=%s, 参数=%s]", req_id, status, device_id, param_name)

        # 如果是已确认状态，需要执行实际的写入
        if status == "CONFIRMED":
            self._execute_confirmed_write(req_id, operation_dict)
        elif status == "CANCELLED":
            self._status_msg_label.setText(f"写操作已取消: {param_name}")
        elif status == "EXECUTED":
            success = operation_dict.get("result", False)
            if success:
                self._status_msg_label.setText(f"✓ 写入成功: {param_name}")
            else:
                error = operation_dict.get("error_message", "未知错误")
                self._status_msg_label.setText(f"✗ 写入失败: {param_name} - {error}")
            orig_req_id = self._undo_req_map.pop(req_id, None)
            if orig_req_id and hasattr(self, "_undo_mgr"):
                if success:
                    self._undo_mgr.mark_undo_success(orig_req_id)
                else:
                    self._undo_mgr.mark_undo_failed(orig_req_id, error)
        elif status == "ABORTED":
            error = operation_dict.get("error_message", "未知错误")
            self._status_msg_label.setText(f"✗ 写入中止: {param_name} - {error}")
            orig_req_id = self._undo_req_map.pop(req_id, None)
            if orig_req_id and hasattr(self, "_undo_mgr"):
                self._undo_mgr.mark_undo_failed(orig_req_id, error)

    def _execute_confirmed_write(self, req_id: str, operation_dict: dict) -> None:
        """
        执行已确认的写操作

        从 operation_dict 中提取参数，
        调用 DeviceConnection.confirm_write() 执行实际写入。

        Args:
            req_id: 请求ID
            operation_dict: 操作详情字典
        """
        device_id = operation_dict.get("device_id", "")
        param_name = operation_dict.get("param_name", "")
        value = operation_dict.get("value")

        # 验证当前选中的设备是否匹配
        if device_id != self._current_device_id:
            logger.warning("设备ID不匹配 [期望=%s, 当前=%s]", device_id, self._current_device_id)
            return

        # 获取设备连接对象
        device = self._device_manager.get_device(device_id)
        if not device or not hasattr(device, "confirm_write"):
            logger.error("设备不支持写操作或设备不存在: %s", device_id)
            self._write_manager.mark_executed(req_id, False, "设备不存在或不支持写操作")
            return

        try:
            # 执行实际的写入
            success = device.confirm_write(param_name, bool(value))

            # 标记操作完成
            self._write_manager.mark_executed(req_id, success)

            # 更新状态栏
            if success:
                self._status_msg_label.setText(f"写入成功 [{param_name}={'ON' if value else 'OFF'}]")
                logger.info("线圈写入成功 [设备=%s, 参数=%s, 值=%s]", device_id, param_name, "ON" if value else "OFF")
            else:
                self._status_msg_label.setText(f"写入失败: {param_name}")
                logger.error("线圈写入失败 [设备=%s, 参数=%s]", device_id, param_name)

        except Exception as e:
            error_msg = f"执行写入异常: {str(e)}"
            logger.error("线圈写入异常 [设备=%s, 参数=%s]: %s", device_id, param_name, str(e), exc_info=True)
            self._write_manager.mark_executed(req_id, False, error_msg)

    def _on_add_device_with_monitor(self) -> None:
        """
        添加设备并自动创建动态监控面板（v3.2 增强版）

        在原有添加设备逻辑基础上，
        成功后自动创建 DynamicMonitorPanel 并加入 Tab 页面。
        """
        from ui.device_type_dialogs import DeviceTypeManager

        logger.debug("Opening add device dialog with monitor panel support")
        try:
            device_type_manager = DeviceTypeManager("device_types.json")
            dialog = AddDeviceDialog(device_type_manager, self)
            if dialog.exec():
                config = dialog.get_device_config()
                device_id = self._device_manager.add_device(config)

                if device_id:
                    # ✅ 创建动态监控面板
                    register_points_config = config.get("register_points", [])
                    if register_points_config:
                        from core.enums.data_type_enum import RegisterPointConfig

                        # 转换为 RegisterPointConfig 对象列表
                        rp_list = []
                        for rp_dict in register_points_config:
                            try:
                                rp = RegisterPointConfig.from_dict(rp_dict)
                                rp_list.append(rp)
                            except Exception as e:
                                logger.warning("转换RegisterPointConfig失败: %s", str(e))

                        # 创建面板
                        monitor_panel = DynamicMonitorPanel(device_id, self)
                        monitor_panel.build_from_config(rp_list)

                        # 连接写请求信号
                        monitor_panel.coil_write_requested.connect(self._on_coil_write_request)

                        # 存储到字典
                        self._monitor_panels[device_id] = monitor_panel

                        # 添加到 Tab 页面
                        tab_text = f"{config.get('name', device_id)}_监控"
                        self._monitor_tabs.addTab(monitor_panel, tab_text)

                        logger.info("动态监控面板已创建 [设备=%s, 数据点=%d]", device_id, len(rp_list))

                    logger.info(LogMessages.DEVICE_ADDED.format(device_id=device_id))
                    self._status_msg_label.setText(f"设备添加成功: {device_id}")

        except Exception as e:
            logger.error("添加设备失败: %s", str(e))
            QMessageBox.warning(self, "添加设备失败", f"无法添加设备: {str(e)}")
            self._status_msg_label.setText(f"添加设备失败: {str(e)}")
